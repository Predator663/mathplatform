"""
Excel Export Engine using openpyxl.
Generates styled workbooks with school header and platform footer.
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.chart import LineChart, BarChart, PieChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.image import Image as XLImage

from .badge_art import badge_png_bytes

# ── Brand colours (openpyxl uses ARGB hex) ────────────────────────────────────
BLUE   = '002563eb'
DARK   = 'FF0a0a0f'
GRAY   = 'FF6b7280'
LIGHT  = 'FFF3F4F6'
GREEN  = 'FF10b981'
AMBER  = 'FFf59e0b'
ROSE   = 'FFf43f5e'
VIOLET = 'FF8b5cf6'
WHITE  = 'FFFFFFFF'
HEADER = 'FF1e3a5f'
SUBHDR = 'FF2d4f7c'


def _fill(hex_color): return PatternFill('solid', fgColor=hex_color)

# ── Formula-injection guard ──────────────────────────────────────────────
# Any cell whose ENTIRE value is user-controlled free text (a student's
# name, a tournament title someone typed in, a challenge label) must never
# be allowed to start with =, +, -, or @ — Excel/Sheets/LibreOffice treat
# a leading one of those as "this cell is a formula" and will evaluate it
# when the file is opened, which is a well-known spreadsheet attack vector
# (CSV/Excel formula injection). Prefixing with a straight quote forces it
# to be stored and displayed as literal text instead.
_FORMULA_TRIGGER_CHARS = ('=', '+', '-', '@', '\t', '\r')


def _safe(value):
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGER_CHARS:
        return "'" + value
    return value


def _font(bold=False, color='FF000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')
def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border(style='thin'):
    s = Side(style=style, color='FFD1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)


def _grade_color_hex(pct):
    # Mirrors pdf_engine._grade_color exactly (and ExamScore.letter_grade
    # bands) so a score is coloured/labelled identically in the PDF and
    # Excel exports and in the app itself.
    if pct >= 75: return GREEN[2:]
    if pct >= 65: return '2563eb'
    if pct >= 45: return AMBER[2:]
    return ROSE[2:]


def _letter_grade_xl(pct):
    if pct >= 75: return 'A'
    if pct >= 65: return 'B'
    if pct >= 45: return 'C'
    if pct >= 30: return 'D'
    return 'F'


def _write_platform_header(ws, school_name, doc_title, doc_subtitle, academic_year, ncols):
    """Write the document header rows at the top of every sheet."""
    # Row 1: Platform name + school
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, f'MathPlatform Analytics  ·  {school_name}  ·  {academic_year}')
    c.font = _font(bold=True, color=WHITE, size=12)
    c.fill = _fill(HEADER[2:])  # strip FF prefix
    c.fill = PatternFill('solid', fgColor=HEADER[2:])
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 22

    # Row 2: Document title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2, 1, doc_title)
    c2.font = _font(bold=True, color=WHITE, size=11)
    c2.fill = PatternFill('solid', fgColor=SUBHDR[2:])
    c2.alignment = _align('center')
    ws.row_dimensions[2].height = 18

    # Row 3: Subtitle / meta
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c3 = ws.cell(3, 1, doc_subtitle)
    c3.font = _font(italic=True, color='FF374151', size=9)
    c3.fill = PatternFill('solid', fgColor='FFE8EFF9')
    c3.alignment = _align('center')
    ws.row_dimensions[3].height = 14

    # Row 4: Generated date
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    c4 = ws.cell(4, 1, f'Generated: {date.today().strftime("%d %B %Y")}')
    c4.font = _font(italic=True, color=GRAY[2:], size=8)
    c4.alignment = _align('right')
    ws.row_dimensions[4].height = 12

    return 5  # next row index


def _freeze_and_autofit(ws, freeze_row, freeze_col, min_widths=None, fixed_cols=None):
    """Freeze panes and auto-size columns to fit their content.

    fixed_cols: optional set of 1-indexed column numbers that should NOT be
    measured against their cell content (e.g. columns holding long, rotated
    header text like exam titles). For those columns, only min_widths is
    used — this stops a single long header from blowing that column out to
    40 chars wide and pushing every other column off-screen/off-page.
    """
    ws.freeze_panes = ws.cell(freeze_row, freeze_col)
    fixed_cols = fixed_cols or set()

    # Cells that anchor a multi-column merge (e.g. the wide "MathPlatform ·
    # School · Year" banner in row 1) hold a full-width string but live in
    # column A. Measuring that string against column A alone used to blow
    # the rank/# column out to ~40 chars wide, overlapping/crowding every
    # column after it. Skip those anchors entirely when sizing.
    merged_anchor_cells = {
        (rng.min_row, rng.min_col)
        for rng in ws.merged_cells.ranges
        if rng.min_col != rng.max_col
    }

    for col in ws.columns:
        col_num = col[0].column
        col_letter = get_column_letter(col_num)
        if col_num in fixed_cols:
            adjusted = (min_widths[col_num - 1]
                        if min_widths and col_num <= len(min_widths) else 10)
            ws.column_dimensions[col_letter].width = adjusted
            continue
        max_len = 0
        for cell in col:
            if (cell.row, cell.column) in merged_anchor_cells:
                continue
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_len + 2, 8), 40)
        if min_widths and col_num <= len(min_widths):
            adjusted = max(adjusted, min_widths[col_num - 1])
        ws.column_dimensions[col_letter].width = adjusted


def _write_badges_sheet(wb, student, school_name, academic_year, badges, tournament_stats=None):
    """'Badges & Prizes' sheet: one row per earned badge with an embedded
    PNG medallion image (real badge art, not just text), plus a KPI strip
    up top for tournament titles/duel wins when available. Mirrors the PDF
    honors section so the Excel and PDF student reports show the same
    prizes."""
    ws = wb.create_sheet('Badges & Prizes')
    ws.sheet_view.showGridLines = False
    ncols = 4
    row = _write_platform_header(
        ws, school_name, _safe(f'{student.full_name} — Badges & Prizes'),
        'Every badge earned for exam performance, quiz consistency, and tournament results.',
        academic_year, ncols,
    )
    row += 1

    ts = tournament_stats or {}
    if any(ts.get(k) for k in ('titles', 'match_wins', 'participations', 'rising_star_count')):
        kpis = [
            ('Tournaments Entered', ts.get('participations') or 0),
            ('Tournament Titles', ts.get('titles') or 0),
            ('Duel Wins', ts.get('match_wins') or 0),
            ('Rising Star Moments', ts.get('rising_star_count') or 0),
        ]
        for j, (label, val) in enumerate(kpis, 1):
            lc = ws.cell(row, j, label)
            lc.font = _font(bold=True, color=GRAY[2:], size=8)
            lc.alignment = _align('center')
            vc = ws.cell(row + 1, j, val)
            vc.font = _font(bold=True, size=13)
            vc.fill = PatternFill('solid', fgColor='FFE8EFF9')
            vc.alignment = _align('center')
            vc.border = _border()
        row += 3

    headers = ['Badge', 'Icon', 'Description', 'Earned']
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill('solid', fgColor=HEADER[2:])
        c.alignment = _align('center')
        c.border = _border()
    ws.row_dimensions[row].height = 18
    header_row = row
    row += 1

    if not badges:
        ws.cell(row, 1, 'No badges earned yet.')
        ws.cell(row, 1).font = _font(italic=True, color=GRAY[2:], size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    else:
        for sb in badges:
            ws.row_dimensions[row].height = 42
            name_c = ws.cell(row, 1, sb.badge.name)
            name_c.font = _font(bold=True, size=9)
            name_c.alignment = _align('left', wrap=True)
            name_c.border = _border()

            # Icon column B: embed the same medallion art used in the PDF.
            img_cell = f'B{row}'
            png_bytes = badge_png_bytes(sb.badge.icon, size=48)
            xl_img = XLImage(io.BytesIO(png_bytes))
            xl_img.width = 34
            xl_img.height = 34
            ws.add_image(xl_img, img_cell)
            ws.cell(row, 2).border = _border()

            desc_c = ws.cell(row, 3, sb.badge.description)
            desc_c.font = _font(size=8.5)
            desc_c.alignment = _align('left', wrap=True)
            desc_c.border = _border()

            date_c = ws.cell(row, 4, sb.awarded_at.strftime('%d %b %Y'))
            date_c.font = _font(size=8.5)
            date_c.alignment = _align('center')
            date_c.border = _border()

            fill = PatternFill('solid', fgColor='FFF8FAFC') if row % 2 == 0 else PatternFill('solid', fgColor=WHITE)
            for col in (1, 3, 4):
                ws.cell(row, col).fill = fill
            row += 1

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 48
    ws.column_dimensions['D'].width = 14
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    return ws


def generate_exam_scores_excel(exam, scores, sort_by='name',
                                school_name='School of Excellence') -> bytes:
    scores_list = list(scores)

    sort_map = {
        'name':       lambda s: s.student.full_name.lower(),
        'score_desc': lambda s: -float(s.score),
        'score_asc':  lambda s: float(s.score),
        'grade':      lambda s: s.letter_grade,
        'student_id': lambda s: s.student.student_id,
    }
    scores_list.sort(key=sort_map.get(sort_by, sort_map['name']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Scores'
    ws.sheet_view.showGridLines = False

    classroom_names = ', '.join(c.name for c in exam.classrooms.all()) or '—'
    present = [s for s in scores_list if not s.is_absent]
    absent_count = len(scores_list) - len(present)
    pcts = [s.percentage for s in present]
    avg = round(sum(pcts)/len(pcts), 1) if pcts else 0
    pass_rate = round(sum(1 for s in present if s.passed)/len(present)*100, 1) if present else 0

    ncols = 9
    next_row = _write_platform_header(
        ws, school_name,
        f'{exam.title} — Score Report',
        f'Type: {exam.get_exam_type_display()}  |  Term: {exam.get_term_display()}  |  '
        f'Date: {exam.exam_date.strftime("%d %b %Y")}  |  Class: {classroom_names}  |  '
        f'Max: {exam.max_score}  |  Pass: {exam.passing_score}',
        exam.academic_year,
        ncols
    )

    # Summary stats row
    next_row += 1
    stats = [('Students', len(scores_list)), ('Present', len(present)), ('Absent', absent_count),
             ('Average', f'{avg}%' if pcts else '—'), ('Pass Rate', f'{pass_rate}%' if present else '—'),
             ('Highest', f'{max(pcts)}%' if pcts else '—'), ('Lowest', f'{min(pcts)}%' if pcts else '—')]
    for j, (label, val) in enumerate(stats, 1):
        lc = ws.cell(next_row, j, label)
        lc.font = _font(bold=True, color='FF6B7280', size=7)
        lc.alignment = _align('center')
        vc = ws.cell(next_row+1, j, val)
        vc.font = _font(bold=True, size=10)
        vc.fill = PatternFill('solid', fgColor='FFE8EFF9')
        vc.alignment = _align('center')
        vc.border = _border()
    next_row += 3

    # Column headers
    col_headers = ['#', 'Student ID', 'Student Name', 'Stream', 'Score', 'Max Score', '% Score', 'Grade', 'Pass?']
    for j, h in enumerate(col_headers, 1):
        c = ws.cell(next_row, j, h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill('solid', fgColor=HEADER[2:])
        c.alignment = _align('center')
        c.border = _border()
    ws.row_dimensions[next_row].height = 18
    next_row += 1
    data_start = next_row

    for rank, s in enumerate(scores_list, 1):
        is_even = rank % 2 == 0
        row_fill = PatternFill('solid', fgColor='FFF8FAFC') if is_even else PatternFill('solid', fgColor=WHITE)
        stream_name = s.student.stream.name if s.student.stream_id else '—'

        if s.is_absent:
            row_data = [rank, s.student.student_id, _safe(s.student.full_name), stream_name, 'ABSENT', exam.max_score, '—', '—', '—']
        else:
            row_data = [rank, s.student.student_id, _safe(s.student.full_name), stream_name,
                        float(s.score), float(exam.max_score), s.percentage/100,
                        s.letter_grade, 'Pass' if s.passed else 'Fail']

        for j, val in enumerate(row_data, 1):
            c = ws.cell(next_row, j, val)
            c.fill = row_fill
            c.border = _border()
            c.alignment = _align('center' if j != 3 else 'left')
            c.font = _font(size=9)

        # Percentage as percentage format
        if not s.is_absent:
            ws.cell(next_row, 7).number_format = '0.0%'
            # Colour pass/fail
            pf_cell = ws.cell(next_row, 9)
            pf_cell.font = _font(bold=True, size=9,
                                  color=GREEN[2:] if s.passed else ROSE[2:])
            # Colour grade
            ws.cell(next_row, 8).font = _font(bold=True, size=9,
                color=GREEN[2:] if s.percentage>=75 else AMBER[2:] if s.percentage>=45 else ROSE[2:])

        next_row += 1

    # Totals row
    ws.cell(next_row, 3, 'CLASS AVERAGE').font = _font(bold=True, size=9)
    ws.cell(next_row, 7, avg/100).number_format = '0.0%'
    ws.cell(next_row, 7).font = _font(bold=True, size=9)
    for j in range(1, ncols+1):
        ws.cell(next_row, j).fill = PatternFill('solid', fgColor='FFE8EFF9')
        ws.cell(next_row, j).border = _border()

    _freeze_and_autofit(ws, data_start, 3)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_view.showGridLines = False
    _write_platform_header(ws2, school_name, f'{exam.title} — Summary',
                            f'Generated: {date.today().strftime("%d %b %Y")}', exam.academic_year, 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_class_report_excel(classroom, students, scores_map, exams,
                                  sort_by='name', school_name='School of Excellence',
                                  top_achievers=None) -> bytes:
    students = list(students)
    exams = list(exams)

    rows_data = []
    for s in students:
        s_scores = scores_map.get(s.id, {})
        pcts = [v for v in s_scores.values() if v is not None]
        avg = round(sum(pcts)/len(pcts), 1) if pcts else None
        rows_data.append((s, s_scores, avg))

    sort_fns = {
        'name':         lambda x: x[0].full_name.lower(),
        'average_desc': lambda x: -(x[2] or 0),
        'average_asc':  lambda x: (x[2] or 0),
        'student_id':   lambda x: x[0].student_id,
    }
    rows_data.sort(key=sort_fns.get(sort_by, sort_fns['name']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Class Report'
    ws.sheet_view.showGridLines = False

    ncols = 4 + len(exams) + 1  # rank, id, name, stream + exams + avg
    _write_platform_header(
        ws, school_name,
        f'{classroom} — Class Performance Report',
        f'Grade: {classroom.grade_level.name}  |  Year: {classroom.academic_year}  |  '
        f'Students: {len(students)}  |  Exams: {len(exams)}',
        classroom.academic_year,
        ncols
    )

    next_row = 6
    col_headers = ['#', 'Student ID', 'Name', 'Stream'] + [e.title for e in exams] + ['AVERAGE']
    # Exam-title columns are rotated 90° so a long exam name doesn't force
    # the column itself to widen (which used to push later columns off the
    # printable page and made scores overlap the header text). The rank/ID/
    # name/stream/average columns stay horizontal since they're short labels.
    exam_col_nums = set(range(5, 5 + len(exams)))
    for j, h in enumerate(col_headers, 1):
        c = ws.cell(next_row, j, h)
        c.font = _font(bold=True, color=WHITE, size=8)
        bg = HEADER[2:] if j <= 4 or j == ncols else SUBHDR[2:]
        c.fill = PatternFill('solid', fgColor=bg)
        if j in exam_col_nums:
            c.alignment = Alignment(horizontal='center', vertical='bottom',
                                     wrap_text=True, text_rotation=90)
        else:
            c.alignment = _align('center', wrap=True)
        c.border = _border()
    # Tall enough to fit rotated exam titles without clipping or overlapping
    # the data rows beneath them.
    ws.row_dimensions[next_row].height = 110
    next_row += 1

    all_avgs = []
    for rank, (s, s_scores, avg) in enumerate(rows_data, 1):
        is_even = rank % 2 == 0
        row_fill = PatternFill('solid', fgColor='FFF8FAFC') if is_even else PatternFill('solid', fgColor=WHITE)

        ws.cell(next_row, 1, rank).fill = row_fill
        ws.cell(next_row, 1).border = _border()
        ws.cell(next_row, 1).alignment = _align('center')

        ws.cell(next_row, 2, s.student_id).fill = row_fill
        ws.cell(next_row, 2).border = _border()
        ws.cell(next_row, 2).alignment = _align('center')
        ws.cell(next_row, 2).font = _font(size=8)

        ws.cell(next_row, 3, _safe(s.full_name)).fill = row_fill
        ws.cell(next_row, 3).border = _border()
        ws.cell(next_row, 3).font = _font(size=8)

        stream_cell = ws.cell(next_row, 4, s.stream.name if s.stream_id else '—')
        stream_cell.fill = row_fill
        stream_cell.border = _border()
        stream_cell.alignment = _align('center')
        stream_cell.font = _font(size=8)

        for ei, e in enumerate(exams, 5):
            pct = s_scores.get(e.id)
            c = ws.cell(next_row, ei, f'{pct}%' if pct is not None else '—')
            c.fill = row_fill
            c.border = _border()
            c.alignment = _align('center')
            c.font = _font(size=8,
                color=GREEN[2:] if pct is not None and pct>=75 else AMBER[2:] if pct is not None and pct>=45 else ROSE[2:] if pct is not None else GRAY[2:])

        avg_c = ws.cell(next_row, ncols, f'{avg}%' if avg is not None else '—')
        avg_c.fill = PatternFill('solid', fgColor='FFE8EFF9')
        avg_c.font = _font(bold=True, size=8,
            color=GREEN[2:] if avg is not None and avg>=75 else AMBER[2:] if avg is not None and avg>=45 else ROSE[2:] if avg is not None else GRAY[2:])
        avg_c.border = _border()
        avg_c.alignment = _align('center')
        if avg is not None: all_avgs.append(avg)
        next_row += 1

    # Class average footer row
    class_avg = round(sum(all_avgs)/len(all_avgs), 1) if all_avgs else None
    ws.cell(next_row, 3, 'CLASS AVERAGE').font = _font(bold=True, size=9)
    avg_footer_c = ws.cell(next_row, ncols, f'{class_avg}%' if class_avg is not None else '—')
    avg_footer_c.font = _font(bold=True, size=9,
        color=(GREEN[2:] if class_avg is not None and class_avg>=75
               else AMBER[2:] if class_avg is not None and class_avg>=45
               else ROSE[2:] if class_avg is not None else GRAY[2:]))
    for j in range(1, ncols+1):
        ws.cell(next_row, j).fill = PatternFill('solid', fgColor='FFD1FAE5')
        ws.cell(next_row, j).border = _border()

    # Exam columns get a fixed narrow width (they hold "87%" style values,
    # not the rotated header text) instead of being auto-sized off the long
    # exam title — that autofit was the source of the column blowout.
    min_widths = [4, 12, 22, 10] + [9] * len(exams) + [10]
    _freeze_and_autofit(ws, 7, 5, min_widths=min_widths, fixed_cols=exam_col_nums)

    # ── Top Achievers sheet (badge leaderboard) ─────────────────────────────
    if top_achievers:
        aws = wb.create_sheet('Top Achievers')
        aws.sheet_view.showGridLines = False
        arow = _write_platform_header(
            aws, school_name, f'{classroom} — Top Achievers',
            'Ranked by total badges earned — exam, quiz, and tournament combined.',
            classroom.academic_year, 5,
        )
        arow += 1
        headers = ['Rank', 'Medal', 'Student', 'Badges Earned', 'Most Recent Badge']
        for j, h in enumerate(headers, 1):
            c = aws.cell(arow, j, h)
            c.font = _font(bold=True, color=WHITE, size=9)
            c.fill = PatternFill('solid', fgColor=HEADER[2:])
            c.alignment = _align('center')
            c.border = _border()
        aws.row_dimensions[arow].height = 18
        header_row = arow
        arow += 1

        for i, row in enumerate(top_achievers[:10], 1):
            aws.row_dimensions[arow].height = 40
            rank_c = aws.cell(arow, 1, f'#{i}')
            rank_c.font = _font(bold=True, size=10)
            rank_c.alignment = _align('center')
            rank_c.border = _border()

            latest = row.get('latest_badge')
            if latest:
                png_bytes = badge_png_bytes(latest.icon, size=48)
                xl_img = XLImage(io.BytesIO(png_bytes))
                xl_img.width = 32
                xl_img.height = 32
                aws.add_image(xl_img, f'B{arow}')
            aws.cell(arow, 2).border = _border()

            name_c = aws.cell(arow, 3, _safe(row['student'].full_name))
            name_c.font = _font(size=9)
            name_c.border = _border()

            count_c = aws.cell(arow, 4, row['badge_count'])
            count_c.font = _font(bold=True, size=10)
            count_c.alignment = _align('center')
            count_c.fill = PatternFill('solid', fgColor='FFE8EFF9')
            count_c.border = _border()

            latest_c = aws.cell(arow, 5, latest.name if latest else '—')
            latest_c.font = _font(size=8.5)
            latest_c.border = _border()

            fill = PatternFill('solid', fgColor='FFF8FAFC') if i % 2 == 0 else PatternFill('solid', fgColor=WHITE)
            for col in (1, 3, 4, 5):
                aws.cell(arow, col).fill = fill
            arow += 1

        aws.column_dimensions['A'].width = 8
        aws.column_dimensions['B'].width = 7
        aws.column_dimensions['C'].width = 26
        aws.column_dimensions['D'].width = 14
        aws.column_dimensions['E'].width = 28
        aws.freeze_panes = aws.cell(header_row + 1, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _std_dev_xl(values):
    if len(values) < 2:
        return 0
    m = sum(values) / len(values)
    return (sum((v - m) ** 2 for v in values) / len(values)) ** 0.5


def generate_student_report_excel(student, scores, topic_data,
                                   school_name='School of Excellence',
                                   trend=None, comparison=None,
                                   badges=None, tournament_stats=None) -> bytes:
    """
    Individual student report, Excel version — mirrors generate_student_report_pdf.

    Sheets: Summary (KPIs + narrative + 4 native charts, each with a proper
    legend), Badges & Prizes (embedded badge medallion images + tournament
    record), Exam History (with classroom-average comparison columns),
    Topic Mastery, Term Breakdown. Chart source data lives on a hidden
    '_ChartData' sheet so the Summary sheet itself stays readable.
    """
    scores = list(scores)
    trend = trend or {}
    comparison = comparison or {}
    class_by_exam = comparison.get('by_exam') or {}
    rank = comparison.get('rank')
    class_size = comparison.get('class_size') or 0
    percentile = comparison.get('percentile')

    present = [s for s in scores if not s.is_absent]
    pcts = [s.percentage for s in present]
    avg = round(sum(pcts) / len(pcts), 1) if pcts else None
    passed_count = sum(1 for s in present if s.passed)
    highest = max(pcts) if pcts else None
    lowest = min(pcts) if pcts else None
    consistency = round(_std_dev_xl(pcts), 1) if len(pcts) > 1 else 0
    trend_label = (trend.get('trend') or 'no_data').replace('_', ' ').capitalize()

    timeline = trend.get('timeline') or [
        {'exam_id': s.exam_id, 'exam_title': s.exam.title,
         'exam_date': s.exam.exam_date.strftime('%Y-%m-%d'), 'percentage': s.percentage}
        for s in present
    ]
    moving_average = trend.get('moving_average') or []

    wb = Workbook()

    # ── Hidden chart-data sheet ──────────────────────────────────────────────
    dws = wb.create_sheet('_ChartData')
    dws.sheet_state = 'hidden'

    # Trend block: columns A-D
    dws.cell(1, 1, 'Exam'); dws.cell(1, 2, 'Score %'); dws.cell(1, 3, 'Moving Avg'); dws.cell(1, 4, 'Class Avg')
    for i, t in enumerate(timeline, 2):
        dws.cell(i, 1, t['exam_date'][5:])
        dws.cell(i, 2, t['percentage'])
        ma = moving_average[i-2] if i-2 < len(moving_average) else None
        dws.cell(i, 3, ma if ma is not None else None)
        dws.cell(i, 4, class_by_exam.get(t.get('exam_id')))
    trend_rows = len(timeline)

    # Topic block: columns F-G
    dws.cell(1, 6, 'Topic'); dws.cell(1, 7, 'Average')
    for i, t in enumerate(topic_data, 2):
        dws.cell(i, 6, t['topic_name'])
        dws.cell(i, 7, t['average'])
    topic_rows = len(topic_data)

    # Grade distribution block: columns I-J
    grade_counts = {}
    for s in present:
        g = s.letter_grade
        grade_counts[g] = grade_counts.get(g, 0) + 1
    order = ['A', 'B', 'C', 'D', 'F']
    grade_labels = [g for g in order if g in grade_counts]
    dws.cell(1, 9, 'Grade'); dws.cell(1, 10, 'Count')
    for i, g in enumerate(grade_labels, 2):
        dws.cell(i, 9, g)
        dws.cell(i, 10, grade_counts[g])
    grade_rows = len(grade_labels)

    # Comparison block: columns L-N (only exams with a classroom average)
    cmp_timeline = [t for t in timeline if t.get('exam_id') in class_by_exam]
    dws.cell(1, 12, 'Exam'); dws.cell(1, 13, 'Student %'); dws.cell(1, 14, 'Class Avg %')
    for i, t in enumerate(cmp_timeline, 2):
        dws.cell(i, 12, t['exam_title'][:20])
        dws.cell(i, 13, t['percentage'])
        dws.cell(i, 14, class_by_exam[t['exam_id']])
    cmp_rows = len(cmp_timeline)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    ncols = 6
    next_row = _write_platform_header(
        ws, school_name, f'Individual Student Report — {student.full_name}',
        f'Student ID: {student.student_id}  |  Class: {student.classroom or "—"}  |  '
        f'Email: {student.email}',
        scores[0].exam.academic_year if scores else '—', ncols,
    )
    next_row += 1

    kpis = [
        ('Exams Taken', len(present)), ('Overall Avg', f'{avg}%' if avg is not None else '—'),
        ('Pass Rate', f'{round(passed_count/len(present)*100, 1)}%' if present else '—'),
        ('Highest', f'{highest}%' if highest is not None else '—'),
        ('Lowest', f'{lowest}%' if lowest is not None else '—'),
        ('Consistency (σ)', f'{consistency} pts'),
    ]
    for j, (label, val) in enumerate(kpis, 1):
        lc = ws.cell(next_row, j, label)
        lc.font = _font(bold=True, color='FF6B7280', size=7)
        lc.alignment = _align('center')
        vc = ws.cell(next_row+1, j, val)
        vc.font = _font(bold=True, size=11)
        vc.fill = PatternFill('solid', fgColor='FFE8EFF9')
        vc.alignment = _align('center')
        vc.border = _border()
    next_row += 2

    kpis2 = [
        ('Trend', trend_label), ('Predicted Grade', _letter_grade_xl(avg) if avg is not None else '—'),
        ('Class Rank', f'{rank} of {class_size}' if rank else '—'),
        ('Percentile', f'Top {round(100 - percentile, 1)}%' if percentile is not None else '—'),
        ('Classmates Compared', str(class_size) if class_size else '—'), ('', ''),
    ]
    for j, (label, val) in enumerate(kpis2, 1):
        if not label:
            continue
        lc = ws.cell(next_row, j, label)
        lc.font = _font(bold=True, color='FF6B7280', size=7)
        lc.alignment = _align('center')
        vc = ws.cell(next_row+1, j, val)
        vc.font = _font(bold=True, size=11)
        vc.fill = PatternFill('solid', fgColor='FFF3F0FF')
        vc.alignment = _align('center')
        vc.border = _border()
    next_row += 3

    # Narrative
    if topic_data:
        sorted_topics = sorted(topic_data, key=lambda t: t['average'], reverse=True)
        strong = [t['topic_name'] for t in sorted_topics if t['average'] >= 70][:3]
        weak = [t['topic_name'] for t in sorted_topics if t['average'] < 50][:3]
        lines = []
        if strong: lines.append(f"Strengths: {', '.join(strong)}.")
        if weak: lines.append(f"Watch areas: {', '.join(weak)}.")
        if not strong and not weak: lines.append('Performance is fairly even across topics.')
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=ncols)
        c = ws.cell(next_row, 1, '  ·  '.join(lines))
        c.font = _font(italic=True, size=9, color='FF374151')
        c.alignment = _align('left', wrap=True)
        ws.row_dimensions[next_row].height = 16
        next_row += 2

    charts_top = next_row

    # Trend chart (Score % / Moving Avg / Class Avg) — legend on by default
    if trend_rows >= 2:
        lc = LineChart()
        lc.title = 'Score Trend Over Time'
        lc.style = 2
        lc.y_axis.title = '%'
        lc.y_axis.scaling.min = 0
        lc.y_axis.scaling.max = 100
        lc.x_axis.title = 'Exam'
        cats = Reference(dws, min_col=1, min_row=2, max_row=1+trend_rows)
        for col, name in [(2, 'Score %'), (3, 'Moving avg (3)'), (4, 'Classroom avg')]:
            has_any = any(
                dws.cell(r, col).value is not None for r in range(2, 2+trend_rows)
            )
            if not has_any:
                continue
            data = Reference(dws, min_col=col, min_row=1, max_row=1+trend_rows)
            lc.add_data(data, titles_from_data=True)
        lc.set_categories(cats)
        lc.width, lc.height = 17, 9
        ws.add_chart(lc, f'A{charts_top}')

    # Topic mastery bar chart — coloured per bar by performance band
    if topic_rows:
        bc = BarChart()
        bc.type = 'col'
        bc.title = 'Topic Mastery'
        bc.y_axis.title = '%'
        bc.y_axis.scaling.min = 0
        bc.y_axis.scaling.max = 100
        data = Reference(dws, min_col=7, min_row=1, max_row=1+topic_rows)
        cats = Reference(dws, min_col=6, min_row=2, max_row=1+topic_rows)
        bc.add_data(data, titles_from_data=True)
        bc.set_categories(cats)
        series = bc.series[0]
        series.data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(
                solidFill=_grade_color_hex(t['average'])))
            for i, t in enumerate(topic_data)
        ]
        bc.legend = None  # per-bar colours aren't a single legend-able series; key given below instead
        bc.width, bc.height = 12, 9
        ws.add_chart(bc, f'H{charts_top}')
        # Manual colour key, since a single-series chart can't show a
        # per-point legend natively in Excel.
        key_row = charts_top + 19
        ws.cell(key_row, 8, 'Colour key:').font = _font(bold=True, size=8)
        for i, (label, hexcolor) in enumerate([
            ('Strong (75%+)', GREEN[2:]), ('Good (65-74%)', '2563eb'),
            ('Fair (45-64%)', AMBER[2:]), ('Needs support (<45%)', ROSE[2:]),
        ]):
            cc = ws.cell(key_row + 1 + i, 8, '  ' + label)
            cc.font = _font(size=8)
            cc.fill = PatternFill('solid', fgColor=hexcolor)

    charts_top2 = charts_top + 20

    # Grade distribution pie chart — legend on by default
    if grade_rows:
        pc = PieChart()
        pc.title = 'Grade Distribution'
        data = Reference(dws, min_col=10, min_row=1, max_row=1+grade_rows)
        cats = Reference(dws, min_col=9, min_row=2, max_row=1+grade_rows)
        pc.add_data(data, titles_from_data=True)
        pc.set_categories(cats)
        grade_hex = {'A': GREEN[2:], 'B': '2563eb', 'C': AMBER[2:], 'D': 'fb923c', 'F': ROSE[2:]}
        pc.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(
                solidFill=grade_hex.get(g, GRAY[2:])))
            for i, g in enumerate(grade_labels)
        ]
        pc.width, pc.height = 12, 9
        ws.add_chart(pc, f'A{charts_top2}')

    # Student vs classroom average, per exam — legend on by default
    if cmp_rows:
        cbc = BarChart()
        cbc.type = 'col'
        cbc.grouping = 'clustered'
        cbc.title = 'Student vs. Classroom Average'
        cbc.y_axis.title = '%'
        cbc.y_axis.scaling.min = 0
        cbc.y_axis.scaling.max = 100
        cats = Reference(dws, min_col=12, min_row=2, max_row=1+cmp_rows)
        data = Reference(dws, min_col=13, max_col=14, min_row=1, max_row=1+cmp_rows)
        cbc.add_data(data, titles_from_data=True)
        cbc.set_categories(cats)
        cbc.series[0].graphicalProperties.solidFill = '2563eb'
        cbc.series[1].graphicalProperties.solidFill = GRAY[2:]
        cbc.width, cbc.height = 12, 9
        ws.add_chart(cbc, f'H{charts_top2}')

    ws.column_dimensions['A'].width = 14
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 12

    # ── Badges & Prizes sheet ────────────────────────────────────────────────
    _write_badges_sheet(
        wb, student, school_name,
        scores[0].exam.academic_year if scores else '—',
        badges or [], tournament_stats,
    )

    # ── Exam History sheet ───────────────────────────────────────────────────
    hws = wb.create_sheet('Exam History')
    hws.sheet_view.showGridLines = False
    has_cmp = bool(class_by_exam)
    hcols = 11 if has_cmp else 8
    _write_platform_header(hws, school_name, _safe(f'{student.full_name} — Exam History'),
                            'Class Avg / vs Class show the classroom average on that same '
                            'exam and the student\'s difference from it.' if has_cmp else
                            'Full record of every exam taken.',
                            scores[0].exam.academic_year if scores else '—', hcols)
    hr = 6
    headers = ['#', 'Exam', 'Type', 'Date', 'Score', '%'] + \
              (['Class Avg', 'vs Class'] if has_cmp else []) + ['Grade', 'Pass?']
    for j, h in enumerate(headers, 1):
        c = hws.cell(hr, j, h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill('solid', fgColor=HEADER[2:])
        c.alignment = _align('center')
        c.border = _border()
    hws.row_dimensions[hr].height = 18
    hr += 1
    for i, s in enumerate(scores, 1):
        row_fill = PatternFill('solid', fgColor='FFF8FAFC') if i % 2 == 0 else PatternFill('solid', fgColor=WHITE)
        vals = [i, s.exam.title, s.exam.get_exam_type_display(), s.exam.exam_date.strftime('%d %b %Y')]
        if s.is_absent:
            vals += ['ABSENT', '—'] + (['—', '—'] if has_cmp else []) + ['—', '—']
        else:
            vals += [float(s.score), s.percentage/100]
            if has_cmp:
                cavg = class_by_exam.get(s.exam_id)
                diff = round(s.percentage - cavg, 1) if cavg is not None else None
                vals += [cavg/100 if cavg is not None else '—', f'{"+" if diff and diff>0 else ""}{diff}%' if diff is not None else '—']
            vals += [s.letter_grade, 'Pass' if s.passed else 'Fail']
        for j, v in enumerate(vals, 1):
            c = hws.cell(hr, j, v)
            c.fill = row_fill
            c.border = _border()
            c.alignment = _align('center' if j != 2 else 'left')
            c.font = _font(size=9)
            if j == 6 and not s.is_absent:
                c.number_format = '0.0%'
            if has_cmp and j == 7 and not s.is_absent:
                c.number_format = '0.0%'
        if not s.is_absent:
            pass_col = 9 if has_cmp else 8
            pc = hws.cell(hr, pass_col)
            pc.font = _font(bold=True, size=9, color=GREEN[2:] if s.passed else ROSE[2:])
            if has_cmp:
                cavg = class_by_exam.get(s.exam_id)
                if cavg is not None:
                    diff_c = hws.cell(hr, 7)
                    diff_c.font = _font(bold=True, size=9,
                                         color=GREEN[2:] if s.percentage >= cavg else ROSE[2:])
        hr += 1
    _freeze_and_autofit(hws, 7, 2, min_widths=[4, 24, 12, 11, 8, 8] + ([9, 9] if has_cmp else []) + [7, 7])

    # ── Topic Mastery sheet ──────────────────────────────────────────────────
    if topic_data:
        tws = wb.create_sheet('Topic Mastery')
        tws.sheet_view.showGridLines = False
        _write_platform_header(tws, school_name, _safe(f'{student.full_name} — Topic Mastery'),
                                'Average performance broken down by topic.',
                                scores[0].exam.academic_year if scores else '—', 7)
        tr = 6
        for j, h in enumerate(['Topic', 'Average %', 'Grade', 'Attempts', 'Highest', 'Lowest', 'Trend'], 1):
            c = tws.cell(tr, j, h)
            c.font = _font(bold=True, color=WHITE, size=9)
            c.fill = PatternFill('solid', fgColor=VIOLET[2:])
            c.alignment = _align('center')
            c.border = _border()
        tr += 1
        for t in topic_data:
            row_fill = PatternFill('solid', fgColor='FFF8FAFC') if tr % 2 == 0 else PatternFill('solid', fgColor=WHITE)
            vals = [t['topic_name'], t['average']/100, _letter_grade_xl(t['average']),
                    t['attempts'], t.get('highest', 0)/100, t.get('lowest', 0)/100, t['trend'].capitalize()]
            for j, v in enumerate(vals, 1):
                c = tws.cell(tr, j, v)
                c.fill = row_fill
                c.border = _border()
                c.alignment = _align('center' if j != 1 else 'left')
                c.font = _font(size=9,
                    color=_grade_color_hex(t['average']) if j == 2 else '000000', bold=(j == 2))
                if j in (2, 5, 6):
                    c.number_format = '0.0%'
            tr += 1
        _freeze_and_autofit(tws, 7, 2, min_widths=[22, 10, 8, 9, 9, 9, 10])

    # ── Term Breakdown sheet ─────────────────────────────────────────────────
    term_groups = {}
    term_labels = {}
    for s in present:
        key = (s.exam.academic_year, s.exam.term)
        term_groups.setdefault(key, []).append(s.percentage)
        term_labels[key] = s.exam.get_term_display()
    if term_groups:
        rws = wb.create_sheet('Term Breakdown')
        rws.sheet_view.showGridLines = False
        _write_platform_header(rws, school_name, _safe(f'{student.full_name} — Term-by-Term Performance'),
                                'Average, highest and lowest score per academic term.',
                                scores[0].exam.academic_year if scores else '—', 6)
        rr = 6
        for j, h in enumerate(['Academic Year', 'Term', 'Exams', 'Average', 'Highest', 'Lowest'], 1):
            c = rws.cell(rr, j, h)
            c.font = _font(bold=True, color=WHITE, size=9)
            c.fill = PatternFill('solid', fgColor=HEADER[2:])
            c.alignment = _align('center')
            c.border = _border()
        rr += 1
        for (year, term), vals in sorted(term_groups.items()):
            row_fill = PatternFill('solid', fgColor='FFF8FAFC') if rr % 2 == 0 else PatternFill('solid', fgColor=WHITE)
            row = [year, term_labels[(year, term)], len(vals),
                   round(sum(vals)/len(vals), 1)/100, max(vals)/100, min(vals)/100]
            for j, v in enumerate(row, 1):
                c = rws.cell(rr, j, v)
                c.fill = row_fill
                c.border = _border()
                c.alignment = _align('center')
                c.font = _font(size=9)
                if j in (4, 5, 6):
                    c.number_format = '0.0%'
            rr += 1
        _freeze_and_autofit(rws, 7, 3, min_widths=[14, 12, 8, 10, 10, 10])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_tournament_dossier_excel(tournament, dossier, analytics,
                                       school_name='School of Excellence') -> bytes:
    """
    Tournament dossier, Excel version — mirrors generate_tournament_dossier_pdf.

    Sheets: Summary (KPIs + champion/rising-star callouts with embedded
    badge medallions + native score-distribution chart), Leaderboard
    (full ranked results), Challenge Log (every duel and its outcome).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    ncols = 5
    row = _write_platform_header(
        ws, school_name, f'Tournament Dossier — {tournament.title}',
        f'{tournament.get_mode_display()}  ·  {tournament.classroom}  ·  Decisive exam: {tournament.exam.title}',
        tournament.exam.academic_year, ncols,
    )
    row += 1

    kpis = [
        ('Status', tournament.get_status_display()),
        ('Entrants', analytics['entrant_count']),
        ('Participation', f"{analytics['participation_rate']}%" if analytics['participation_rate'] is not None else '—'),
        ('Pass Rate', f"{analytics['pass_rate']}%" if analytics['pass_rate'] is not None else '—'),
        ('Entrant Avg', f"{analytics['entrant_average']}%" if analytics['entrant_average'] is not None else '—'),
    ]
    for j, (label, val) in enumerate(kpis, 1):
        lc = ws.cell(row, j, label)
        lc.font = _font(bold=True, color=GRAY[2:], size=8)
        lc.alignment = _align('center')
        vc = ws.cell(row + 1, j, val)
        vc.font = _font(bold=True, size=12)
        vc.fill = PatternFill('solid', fgColor='FFE8EFF9')
        vc.alignment = _align('center')
        vc.border = _border()
    row += 3

    # Champion callout with embedded medallion
    champion = dossier.get('champion')
    if champion:
        ws.cell(row, 1, 'CHAMPION').font = _font(bold=True, color=AMBER[2:], size=9)
        row += 1
        png_bytes = badge_png_bytes('trophy', size=56)
        xl_img = XLImage(io.BytesIO(png_bytes))
        xl_img.width = 36
        xl_img.height = 36
        ws.add_image(xl_img, f'A{row}')
        c = ws.cell(row, 2, _safe(f'{champion.entry.display_name} — {champion.score_percentage}%'))
        c.font = _font(bold=True, size=11)
        ws.row_dimensions[row].height = 30
        row += 2

    rising_stars = dossier.get('rising_stars') or []
    if rising_stars:
        ws.cell(row, 1, 'RISING STARS').font = _font(bold=True, color=VIOLET[2:], size=9)
        row += 1
        for r in rising_stars[:5]:
            png_bytes = badge_png_bytes('sparkles', size=40)
            xl_img = XLImage(io.BytesIO(png_bytes))
            xl_img.width = 26
            xl_img.height = 26
            ws.add_image(xl_img, f'A{row}')
            c = ws.cell(row, 2, _safe(f'{r.entry.display_name} — {r.score_percentage}% (+{r.delta:.1f} on prior avg)'))
            c.font = _font(size=9.5)
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1

    # Score distribution native chart
    dist = analytics.get('score_distribution') or []
    if dist:
        dws = wb.create_sheet('_ChartData')
        dws.sheet_state = 'hidden'
        dws.cell(1, 1, 'Band')
        dws.cell(1, 2, 'Count')
        for i, b in enumerate(dist, 2):
            dws.cell(i, 1, b['band'])
            dws.cell(i, 2, b['count'])

        bc = BarChart()
        bc.title = 'Score Distribution'
        bc.y_axis.title = 'Entrants'
        data = Reference(dws, min_col=2, min_row=1, max_row=1 + len(dist))
        cats = Reference(dws, min_col=1, min_row=2, max_row=1 + len(dist))
        bc.add_data(data, titles_from_data=True)
        bc.set_categories(cats)
        bc.legend = None
        bc.width, bc.height = 14, 8
        ws.add_chart(bc, f'A{row}')
        row += 18

    ws.column_dimensions['A'].width = 14
    for col in 'BCDE':
        ws.column_dimensions[col].width = 16

    # ── Leaderboard sheet ────────────────────────────────────────────────────
    lws = wb.create_sheet('Leaderboard')
    lws.sheet_view.showGridLines = False
    lrow = _write_platform_header(
        lws, school_name, _safe(f'{tournament.title} — Final Leaderboard'), 'Ranked by decisive-exam score.',
        tournament.exam.academic_year, 6,
    )
    lrow += 1
    headers = ['Rank', 'Entrant', 'Score', 'Prior Avg', 'Change', 'Flags']
    for j, h in enumerate(headers, 1):
        c = lws.cell(lrow, j, h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill('solid', fgColor=HEADER[2:])
        c.alignment = _align('center')
        c.border = _border()
    header_row = lrow
    lrow += 1
    for i, r in enumerate(dossier['results']):
        flags = []
        if r.is_champion: flags.append('Champion')
        if r.is_rising_star: flags.append('Rising Star')
        if r.is_absent: flags.append('Absent')
        row_vals = [
            f'#{r.rank}' if r.rank else '—', _safe(r.entry.display_name),
            f'{r.score_percentage}%' if r.score_percentage is not None else '—',
            f'{r.prior_average}%' if r.prior_average is not None else '—',
            f'{r.delta:+.1f}' if r.delta is not None else '—',
            ', '.join(flags) or '—',
        ]
        fill = PatternFill('solid', fgColor='FFFEF3C7') if r.is_champion else (
            PatternFill('solid', fgColor='FFF8FAFC') if i % 2 else PatternFill('solid', fgColor=WHITE))
        for j, v in enumerate(row_vals, 1):
            c = lws.cell(lrow, j, v)
            c.font = _font(bold=(j == 1 or r.is_champion), size=9)
            c.alignment = _align('center' if j in (1, 3, 4, 5) else 'left')
            c.fill = fill
            c.border = _border()
        lrow += 1
    _freeze_and_autofit(lws, header_row + 1, 1, min_widths=[8, 24, 10, 12, 8, 22])

    # ── Challenge Log sheet ──────────────────────────────────────────────────
    challenges = list(tournament.challenges.prefetch_related('entries__student', 'entries__stream').select_related('winner').all())
    if challenges:
        cws = wb.create_sheet('Challenge Log')
        cws.sheet_view.showGridLines = False
        crow = _write_platform_header(
            cws, school_name, _safe(f'{tournament.title} — Challenge Log'), 'Every declared duel and its outcome.',
            tournament.exam.academic_year, 4,
        )
        crow += 1
        headers = ['Duel', 'Combatants', 'Result', 'Status']
        for j, h in enumerate(headers, 1):
            c = cws.cell(crow, j, h)
            c.font = _font(bold=True, color=WHITE, size=9)
            c.fill = PatternFill('solid', fgColor=HEADER[2:])
            c.alignment = _align('center')
            c.border = _border()
        chdr_row = crow
        crow += 1
        for i, ch in enumerate(challenges):
            names = ' vs '.join(_safe(e.display_name) for e in ch.entries.all())
            result = _safe(ch.winner.display_name) if ch.winner else ('Tied' if ch.is_tie else '—')
            row_vals = [_safe(ch.label) if ch.label else f'Duel #{ch.id}', names, result, ch.get_status_display()]
            fill = PatternFill('solid', fgColor='FFF8FAFC') if i % 2 else PatternFill('solid', fgColor=WHITE)
            for j, v in enumerate(row_vals, 1):
                c = cws.cell(crow, j, v)
                c.font = _font(size=9)
                c.fill = fill
                c.border = _border()
            crow += 1
        _freeze_and_autofit(cws, chdr_row + 1, 1, min_widths=[22, 40, 20, 16])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_hall_of_fame_excel(hof, *, scope_label='All Classrooms',
                                 school_name='School of Excellence', academic_year='') -> bytes:
    """Hall of Fame, Excel version — mirrors generate_hall_of_fame_pdf.
    One sheet: reigning champions, top-tier standings, and the most-
    promoted leaderboard, stacked with clear section headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hall of Fame'
    ncols = 6

    row = _write_platform_header(ws, school_name, 'League Hall of Fame', f'Scope: {scope_label}', academic_year, ncols)
    row += 1

    top_tier = hof.get('top_tier', [])
    champions = hof.get('season_champions', [])
    most_promoted = hof.get('most_promoted', [])

    def section_title(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1, text)
        c.font = _font(bold=True, color=WHITE, size=11)
        c.fill = PatternFill('solid', fgColor=SUBHDR[2:])
        c.alignment = _align('left')
        row += 1

    def header_row(cols):
        nonlocal row
        for i, label in enumerate(cols, 1):
            c = ws.cell(row, i, label)
            c.font = _font(bold=True, color=WHITE, size=9)
            c.fill = PatternFill('solid', fgColor=HEADER[2:])
            c.alignment = _align('center')
        row += 1

    # ── Reigning Champions ──────────────────────────────────────────────
    section_title('Reigning Champions')
    header_row(['Student', 'Season', 'Classroom', 'Band', 'Score', ''])
    for c in champions:
        ws.cell(row, 1, _safe(c['student_name']))
        ws.cell(row, 2, _safe(c['season_title']))
        ws.cell(row, 3, _safe(c['classroom']))
        ws.cell(row, 4, _safe(c['group_name']))
        sc = ws.cell(row, 5, c['score'])
        sc.alignment = _align('center')
        row += 1
    if not champions:
        ws.cell(row, 1, 'No seasons in scope yet.')
        row += 1
    row += 1

    # ── Top-Tier Standings ───────────────────────────────────────────────
    section_title('Top-Tier Standings')
    header_row(['#', 'Student', 'Classroom', 'Season', 'Band', 'Score'])
    for i, r in enumerate(top_tier, 1):
        ws.cell(row, 1, i).alignment = _align('center')
        ws.cell(row, 2, _safe(r['student_name']))
        ws.cell(row, 3, _safe(r['classroom']))
        ws.cell(row, 4, _safe(r['season_title']))
        ws.cell(row, 5, _safe(r['group_name']))
        ws.cell(row, 6, r['score']).alignment = _align('center')
        row += 1
    if not top_tier:
        ws.cell(row, 1, 'No top-tier members in scope yet.')
        row += 1
    row += 1

    # ── Most Promoted ────────────────────────────────────────────────────
    section_title('Most Promoted')
    header_row(['#', 'Student', 'Total Promotions', '', '', ''])
    for i, r in enumerate(most_promoted, 1):
        ws.cell(row, 1, i).alignment = _align('center')
        ws.cell(row, 2, _safe(r['student_name']))
        ws.cell(row, 3, r['promotion_count']).alignment = _align('center')
        row += 1
    if not most_promoted:
        ws.cell(row, 1, 'No promotions recorded yet.')
        row += 1

    _freeze_and_autofit(ws, freeze_row=6, freeze_col=1, min_widths=[6, 24, 18, 20, 16, 10])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

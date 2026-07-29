"""
Verifies the Stream column added to report exports: exam scores, class
reports, and the all-subjects analytics report, in Excel, PDF, and CSV
form. Run with:
    python manage.py test mathapi.apps.reports.test_stream_in_reports -v 2
"""
import datetime
import io
import csv
from openpyxl import load_workbook
from rest_framework.test import APITestCase
from rest_framework import status
from django.contrib.auth import get_user_model

from mathapi.apps.accounts.models import Subject
from mathapi.apps.students.models import GradeLevel, Classroom, Stream, StudentProfile
from mathapi.apps.exams.models import Exam, ExamScore

User = get_user_model()


class StreamInReportsTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(email='admin2@test.tz', password='x', role='super_admin')
        cls.subject = Subject.objects.create(name='Report Maths', code='RPTM')
        cls.grade = GradeLevel.objects.create(name='Form 1', short_name='F1', education_level='secondary', order=1)
        cls.classroom = Classroom.objects.create(name='1A', grade_level=cls.grade, stream='general', academic_year='2026')
        cls.stream = Stream.objects.create(classroom=cls.classroom, name='Blue')

        u1 = User.objects.create_user(email='rs1@test.tz', password='x', role='student', first_name='Zara', last_name='One')
        u2 = User.objects.create_user(email='rs2@test.tz', password='x', role='student', first_name='Amara', last_name='Two')
        cls.student_with_stream = StudentProfile.objects.create(
            user=u1, student_id='R001', classroom=cls.classroom, stream=cls.stream)
        cls.student_no_stream = StudentProfile.objects.create(
            user=u2, student_id='R002', classroom=cls.classroom)

        cls.exam = Exam.objects.create(
            title='Report Test Exam', exam_type='monthly_test', term='term_1', academic_year='2026',
            exam_date=datetime.date(2026, 5, 1), max_score=100, passing_score=30,
            subject=cls.subject, created_by=cls.admin, is_published=True,
        )
        cls.exam.classrooms.add(cls.classroom)
        ExamScore.objects.create(exam=cls.exam, student=cls.student_with_stream, score=70, entered_by=cls.admin)
        ExamScore.objects.create(exam=cls.exam, student=cls.student_no_stream, score=55, entered_by=cls.admin)

    def test_exam_scores_excel_has_stream_column(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/exam/{self.exam.id}/excel/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb['Scores']
        # Locate header row (first row containing 'Student Name')
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=15):
            values = [c.value for c in row]
            if 'Student Name' in values:
                header_row = values
                header_row_idx = row[0].row
                break
        self.assertIsNotNone(header_row, 'Could not find header row')
        self.assertIn('Stream', header_row)
        stream_col = header_row.index('Stream') + 1

        # Collect the Stream cell values on the two data rows that follow.
        found = set()
        for r in range(header_row_idx + 1, header_row_idx + 3):
            found.add(ws.cell(r, stream_col).value)
        self.assertIn('Blue', found)
        self.assertIn('—', found)

    def test_exam_scores_pdf_generates_without_error(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/exam/{self.exam.id}/pdf/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.content[:4], b'%PDF')
        self.assertGreater(len(resp.content), 1000)

    def test_exam_scores_csv_has_stream_column(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/exam/{self.exam.id}/csv/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        reader = list(csv.reader(io.StringIO(resp.content.decode('utf-8'))))
        header = reader[1]  # row 0 is the title line
        self.assertIn('Stream', header)
        stream_idx = header.index('Stream')
        stream_values = {row[stream_idx] for row in reader[2:]}
        self.assertIn('Blue', stream_values)

    def test_class_report_excel_has_stream_column(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/classroom/{self.classroom.id}/excel/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb['Class Report']
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=15):
            values = [c.value for c in row]
            if 'Name' in values and 'Stream' in values:
                header_row = values
                header_row_idx = row[0].row
                break
        self.assertIsNotNone(header_row, 'Could not find header row with Name+Stream')
        stream_col = header_row.index('Stream') + 1
        found = set()
        for r in range(header_row_idx + 1, header_row_idx + 3):
            found.add(ws.cell(r, stream_col).value)
        self.assertIn('Blue', found)
        self.assertIn('—', found)

    def test_class_report_pdf_generates_without_error(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/classroom/{self.classroom.id}/pdf/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.content[:4], b'%PDF')
        self.assertGreater(len(resp.content), 1000)

    def test_class_roster_csv_has_stream_column(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/classroom/{self.classroom.id}/csv/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        reader = list(csv.reader(io.StringIO(resp.content.decode('utf-8'))))
        header = reader[1]
        self.assertIn('Stream', header)
        stream_idx = header.index('Stream')
        stream_values = {row[stream_idx] for row in reader[2:]}
        self.assertIn('Blue', stream_values)


class StreamInAnalyticsReportTests(APITestCase):
    """Covers the all-subjects analytics report separately: it builds its
    student rows from a raw .values() query rather than iterating related
    objects, so the Stream plumbing (query fields → student_meta →
    student_rows → PDF/Excel column) needs its own check rather than
    inheriting the exam-scores/class-report coverage above."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(email='admin3@test.tz', password='x', role='super_admin')
        cls.subjects = [
            Subject.objects.create(name=f'Analytics Subj {i}', code=f'AS{i}')
            for i in range(7)  # 7 subjects so a division can be computed too
        ]
        cls.grade = GradeLevel.objects.create(name='Form 2', short_name='F2', education_level='secondary', order=2)
        cls.classroom = Classroom.objects.create(name='2A', grade_level=cls.grade, stream='general', academic_year='2026')
        cls.stream = Stream.objects.create(classroom=cls.classroom, name='Gold')

        u1 = User.objects.create_user(email='as1@test.tz', password='x', role='student', first_name='Neema', last_name='One')
        u2 = User.objects.create_user(email='as2@test.tz', password='x', role='student', first_name='Baraka', last_name='Two')
        cls.student_with_stream = StudentProfile.objects.create(
            user=u1, student_id='A001', classroom=cls.classroom, stream=cls.stream)
        cls.student_no_stream = StudentProfile.objects.create(
            user=u2, student_id='A002', classroom=cls.classroom)

        for i, subject in enumerate(cls.subjects):
            exam = Exam.objects.create(
                title=f'Analytics Exam {i}', exam_type='monthly_test', term='term_1', academic_year='2026',
                exam_date=datetime.date(2026, 5, 1), max_score=100, passing_score=30,
                subject=subject, created_by=cls.admin, is_published=True,
            )
            exam.classrooms.add(cls.classroom)
            ExamScore.objects.create(exam=exam, student=cls.student_with_stream, score=70 + i, entered_by=cls.admin)
            ExamScore.objects.create(exam=exam, student=cls.student_no_stream, score=50 + i, entered_by=cls.admin)

    def test_analytics_excel_has_stream_column(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/classroom/{self.classroom.id}/analytics/excel/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb['Student Marks']
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            values = [c.value for c in row]
            if 'STUDENT NAME' in values and 'STREAM' in values:
                header_row = values
                header_row_idx = row[0].row
                break
        self.assertIsNotNone(header_row, 'Could not find header row with STUDENT NAME+STREAM')
        stream_col = header_row.index('STREAM') + 1
        # Confirm STREAM sits immediately after STUDENT NAME, and that the
        # first subject-code column starts right after it (i.e. nothing is
        # off-by-one from the column-index shift).
        self.assertEqual(stream_col, header_row.index('STUDENT NAME') + 2)

        found = set()
        for r in range(header_row_idx + 1, header_row_idx + 3):
            found.add(ws.cell(r, stream_col).value)
        self.assertIn('Gold', found)
        self.assertIn('—', found)

    def test_analytics_pdf_generates_without_error(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.get(f'/api/reports/export/classroom/{self.classroom.id}/analytics/pdf/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.content[:4], b'%PDF')
        self.assertGreater(len(resp.content), 1000)

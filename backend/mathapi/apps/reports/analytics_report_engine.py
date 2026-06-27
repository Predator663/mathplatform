"""
Analytics Report Engine — All-Subjects Summary Report
Generates NECTA-style subject summary tables for PDF and Excel.

REPORT STRUCTURE (2 sections):
  PAGE 1 — STUDENT MARKS TABLE
    SN | REG NO | STUDENT NAME | [SUBJ score/grade] x N | TOTAL | AVERAGE | GRADE | DIV | POS
  PAGE 2 — SUMMARY
    Subject Summary table
    Division Summary  +  Classwise Competency
    List of 10 Best   +  List of 10 Worst
"""
import io
from datetime import date
from collections import defaultdict

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED GRADE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _letter_grade(pct):
    if pct is None: return '-'
    if pct >= 75:   return 'A'
    if pct >= 65:   return 'B'
    if pct >= 45:   return 'C'
    if pct >= 30:   return 'D'
    return 'F'

def _gpa(pct):
    if pct is None: return None
    if pct >= 75:   return 1.0
    if pct >= 65:   return 2.0
    if pct >= 45:   return 3.0
    if pct >= 30:   return 4.0
    return 5.0

def _competency(gpa):
    if gpa is None:  return '-'
    if gpa <= 1.5:   return 'EXCELLENT'
    if gpa <= 2.0:   return 'VERY GOOD'
    if gpa <= 2.5:   return 'GOOD'
    if gpa <= 3.5:   return 'SATISFACTORY'
    if gpa <= 4.0:   return 'AVERAGE'
    return 'FAIL'

def _division(avg_gpa):
    if avg_gpa is None: return '0'
    if avg_gpa <= 1.6:  return 'I'
    if avg_gpa <= 2.0:  return 'II'
    if avg_gpa <= 2.5:  return 'III'
    if avg_gpa <= 3.5:  return 'IV'
    return '0'


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_analytics_report_data(classroom_id, academic_year=None, term=None,
                                 subject_id=None, created_by_id=None, exam_id=None):
    from django.db.models import FloatField, ExpressionWrapper, F, Q
    from mathapi.apps.students.models import StudentProfile, Classroom
    from mathapi.apps.exams.models import Exam, ExamScore

    classroom = Classroom.objects.select_related('grade_level').get(id=classroom_id)

    pct_expr = ExpressionWrapper(
        F('score') * 100.0 / F('exam__max_score'),
        output_field=FloatField(),
    )

    # ── Scope filter ────────────────────────────────────────────────────────
    base_filter = Q(
        student__classroom_id=classroom_id,
        is_absent=False,
        exam__max_score__gt=0,
        exam__is_deleted=False,
    )
    if exam_id:
        base_filter &= Q(exam_id=exam_id)
    else:
        if academic_year:
            base_filter &= Q(exam__academic_year=academic_year)
        if term:
            base_filter &= Q(exam__term=term)
        if created_by_id:
            base_filter &= Q(exam__created_by_id=created_by_id)
        if subject_id:
            base_filter &= Q(exam__subject_id=subject_id)

    # ── Pull all scores ─────────────────────────────────────────────────────
    scores_qs = list(
        ExamScore.objects
        .filter(base_filter)
        .annotate(pct=pct_expr)
        .values(
            'student_id',
            'student__user__first_name',
            'student__user__last_name',
            'student__index_number',
            'exam__subject_id',
            'exam__subject__name',
            'exam__subject__code',
            'score',
            'exam__max_score',
            'pct',
        )
        .order_by('student__user__last_name', 'student__user__first_name')
    )

    # ── Aggregate per (student, subject) ────────────────────────────────────
    # Store lists of (score, max_score, pct) per (student_id, subject_id)
    student_subject_raw: dict = defaultdict(list)   # → [(score, max_score, pct), ...]
    student_meta: dict = {}
    subject_meta: dict = {}

    for row in scores_qs:
        sid  = row['student_id']
        subj = row['exam__subject_id']
        pct  = row['pct']
        if pct is None:
            continue
        student_subject_raw[(sid, subj)].append((
            float(row['score']),
            float(row['exam__max_score']),
            float(pct),
        ))
        if sid not in student_meta:
            fn = row['student__user__first_name'] or ''
            ln = row['student__user__last_name'] or ''
            student_meta[sid] = {
                'name': f'{fn} {ln}'.strip() or f'Student {sid}',
                'index': row['student__index_number'] or '',
            }
        if subj and subj not in subject_meta:
            subject_meta[subj] = {
                'name': row['exam__subject__name'] or f'Subject {subj}',
                'code': row['exam__subject__code'] or '',
            }

    if not student_meta:
        raise ValueError(
            'No scored exams found for this classroom with the selected filters. '
            'Ensure exams have been entered with marks for students in this classroom.'
        )

    # avg pct per (student, subject)
    avg_ss: dict = {}
    for (sid, subj), entries in student_subject_raw.items():
        total_score = sum(e[0] for e in entries)
        total_max   = sum(e[1] for e in entries)
        avg_pct     = (total_score / total_max * 100) if total_max else 0.0
        avg_ss[(sid, subj)] = round(avg_pct, 1)

    # ordered subject list (by code/name)
    ordered_subjects = sorted(
        subject_meta.items(),
        key=lambda x: x[1]['code'] or x[1]['name']
    )  # [(subj_id, {name, code}), ...]

    # ── Per-student row data (for marks table) ──────────────────────────────
    student_rows = []
    for sid, smeta in sorted(student_meta.items(),
                              key=lambda x: x[1]['name']):
        subj_scores = {}  # subj_id → {pct, grade}
        for subj_id, _ in ordered_subjects:
            pct = avg_ss.get((sid, subj_id))
            subj_scores[subj_id] = {
                'pct':   pct,
                'grade': _letter_grade(pct) if pct is not None else '-',
            }
        pcts_all = [v for v in avg_ss.values()
                    if any(k == (sid, _) for k, _ in [(k, None) for k in avg_ss if k[0] == sid])]
        # collect all pcts for this student
        student_pcts = [avg_ss[(sid, subj_id)]
                        for subj_id, _ in ordered_subjects
                        if (sid, subj_id) in avg_ss]
        if not student_pcts:
            continue
        avg_pct = round(sum(student_pcts) / len(student_pcts), 1)
        avg_gpa = round(sum(_gpa(p) for p in student_pcts) / len(student_pcts), 4)
        student_rows.append({
            'student_id':   sid,
            'name':         smeta['name'],
            'index':        smeta['index'],
            'subj_scores':  subj_scores,        # subj_id → {pct, grade}
            'total_pct':    round(sum(student_pcts), 1),
            'average':      avg_pct,
            'grade':        _letter_grade(avg_pct),
            'gpa':          avg_gpa,
            'division':     _division(avg_gpa),
        })

    # Assign positions (rank by average descending)
    student_rows.sort(key=lambda r: -r['average'])
    for i, r in enumerate(student_rows):
        r['position'] = i + 1
    # Re-sort alphabetically for display
    student_rows.sort(key=lambda r: r['name'])

    # ── Subject summary ─────────────────────────────────────────────────────
    subject_rows = []
    for subj_id, smeta in ordered_subjects:
        pcts = [avg_ss[(sid, subj_id)]
                for sid in student_meta
                if (sid, subj_id) in avg_ss]
        if not pcts:
            continue
        grades = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        for p in pcts:
            g = _letter_grade(p)
            if g in grades:
                grades[g] += 1
        avg      = round(sum(pcts) / len(pcts), 2)
        gpa_val  = round(sum(_gpa(p) for p in pcts) / len(pcts), 4)
        subject_rows.append({
            'subject':    smeta['code'] or smeta['name'],
            'full_name':  smeta['name'],
            'A': grades['A'], 'B': grades['B'], 'C': grades['C'],
            'D': grades['D'], 'F': grades['F'],
            'seats':      len(pcts),
            'average':    avg,
            'grade':      _letter_grade(avg),
            'gpa':        round(gpa_val, 4),
            'competency': _competency(gpa_val),
        })

    # ── Per-student overall ─────────────────────────────────────────────────
    student_overall = {r['student_id']: r for r in student_rows}

    # ── Division summary ────────────────────────────────────────────────────
    div_map: dict = defaultdict(int)
    for r in student_rows:
        div_map[r['division']] += 1

    division_summary = {}
    for sex in ('M', 'F', 'TOTAL'):
        row = {}
        for div in ('I', 'II', 'III', 'IV', '0', 'INC'):
            row[div] = div_map.get(div, 0) if sex in ('F', 'TOTAL') else 0
        division_summary[sex] = row

    # ── Classwise competency ────────────────────────────────────────────────
    all_avgs = [r['average'] for r in student_rows]
    all_gpas = [r['gpa'] for r in student_rows]
    class_avg = round(sum(all_avgs) / len(all_avgs), 1) if all_avgs else 0.0
    class_gpa = round(sum(all_gpas) / len(all_gpas), 4) if all_gpas else 0.0

    classwise = {
        'overall_average': class_avg,
        'grade':           _letter_grade(class_avg),
        'gpa':             class_gpa,
        'competency':      _competency(class_gpa),
    }

    # ── Best / worst 10 ────────────────────────────────────────────────────
    ranked = sorted(student_rows, key=lambda r: -r['average'])
    best_10 = [
        {'sn': i+1, 'name': r['name'], 'average': r['average'],
         'grade': r['grade'], 'position': r['position']}
        for i, r in enumerate(ranked[:10])
    ]
    worst_10 = [
        {'sn': i+1, 'name': r['name'], 'average': r['average'],
         'grade': r['grade'], 'position': r['position']}
        for i, r in enumerate(reversed(ranked[-10:]))
    ]

    return {
        'classroom': {
            'name':          classroom.name,
            'grade':         classroom.grade_level.name if classroom.grade_level_id else '',
            'academic_year': academic_year or classroom.academic_year,
            'term':          term or 'All Terms',
        },
        'ordered_subjects': ordered_subjects,    # [(subj_id, {name, code}), ...]
        'student_rows':     student_rows,         # full per-student data
        'subject_rows':     subject_rows,         # aggregated subject summary
        'division_summary': division_summary,
        'classwise':        classwise,
        'best_students':    best_10,
        'worst_students':   worst_10,
        'total_students':   len(student_rows),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF COLOURS & HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

C_DARK   = colors.HexColor('#0a0a0f')
C_BLUE   = colors.HexColor('#1e3a5f')
C_BLUE2  = colors.HexColor('#2d4f7c')
C_LGRAY  = colors.HexColor('#f3f4f6')
C_MGRAY  = colors.HexColor('#d1d5db')
C_GREEN  = colors.HexColor('#10b981')
C_AMBER  = colors.HexColor('#f59e0b')
C_RED    = colors.HexColor('#f43f5e')
C_HEADER = colors.HexColor('#e8f0fe')
WHITE    = colors.white
BLACK    = colors.black

GRADE_COLORS = {
    'A': C_GREEN,
    'B': colors.HexColor('#2563eb'),
    'C': C_AMBER,
    'D': colors.HexColor('#f97316'),
    'F': C_RED,
}

def _ps(name, **kw):
    d = dict(fontName='Helvetica', fontSize=8, textColor=C_DARK,
             spaceAfter=0, spaceBefore=0, leading=10)
    d.update(kw)
    return ParagraphStyle(name, **d)

def _th(text, size=7.5, bold=True, align=TA_CENTER, color=WHITE):
    return Paragraph(text,
        _ps('th', fontName='Helvetica-Bold' if bold else 'Helvetica',
            fontSize=size, alignment=align, textColor=color))

def _td(text, size=7.5, bold=False, align=TA_CENTER, color=C_DARK):
    return Paragraph(str(text) if text is not None else '',
        _ps('td', fontName='Helvetica-Bold' if bold else 'Helvetica',
            fontSize=size, alignment=align, textColor=color))

def _ts_base(nrows):
    return TableStyle([
        ('FONTNAME',       (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE',       (0, 0), (-1, -1), 7.5),
        ('GRID',           (0, 0), (-1, -1), 0.4, C_MGRAY),
        ('BACKGROUND',     (0, 0), (-1, 0),  C_BLUE),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  WHITE),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, nrows-1), [C_LGRAY, WHITE]),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',     (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 2),
    ])


def _header_block(school_name, classroom_info, subtitle=''):
    """Returns (header_table, subtitle_table)."""
    cls = classroom_info
    title = school_name.upper()
    sub   = subtitle or (
        f"ALL-SUBJECTS ANALYTICS REPORT  ·  {cls['name']}  ·  "
        f"{cls['academic_year']}  ·  {str(cls['term']).replace('_',' ').upper()}"
    )
    return title, sub


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF — PAGE 1: STUDENT MARKS TABLE
# ═══════════════════════════════════════════════════════════════════════════════

def _build_marks_table(data, avail_w):
    """Build the student-by-subject marks table flowable."""
    ordered_subjects = data['ordered_subjects']   # [(id, {name, code}), ...]
    student_rows     = data['student_rows']

    n_subj = len(ordered_subjects)

    # ── Column widths ────────────────────────────────────────────────────────
    # Fixed cols: SN(0.5) + REG(2.0) + NAME(4.0) + ... + AVG(1.4) + GRD(1.0) + DIV(0.7) + POS(0.8)
    W_SN   = 0.55 * cm
    W_REG  = 2.00 * cm
    W_NAME = 4.00 * cm
    W_AVG  = 1.40 * cm
    W_GRD  = 0.90 * cm
    W_DIV  = 0.75 * cm
    W_POS  = 0.75 * cm
    fixed  = W_SN + W_REG + W_NAME + W_AVG + W_GRD + W_DIV + W_POS

    # Each subject gets two sub-columns: score% + grade
    remaining = avail_w - fixed
    W_SUBJ_PAIR = max(1.6 * cm, remaining / max(n_subj, 1))
    W_PCT  = W_SUBJ_PAIR * 0.60
    W_GRD2 = W_SUBJ_PAIR * 0.40

    col_widths = [W_SN, W_REG, W_NAME]
    for _ in ordered_subjects:
        col_widths += [W_PCT, W_GRD2]
    col_widths += [W_AVG, W_GRD, W_DIV, W_POS]

    # ── Header row ───────────────────────────────────────────────────────────
    hdr = [
        _th('SN'),
        _th('REG NO'),
        _th('STUDENT NAME', align=TA_LEFT),
    ]
    for _, smeta in ordered_subjects:
        label = smeta['code'] or smeta['name'][:6]
        hdr += [_th(label), _th('GR')]
    hdr += [_th('AVG'), _th('GR'), _th('DIV'), _th('POS')]

    rows = [hdr]

    for i, sr in enumerate(student_rows):
        bg = C_LGRAY if i % 2 == 0 else WHITE
        row_cells = [
            _td(i + 1),
            _td(sr['index'] or '-'),
            _td(sr['name'], align=TA_LEFT, bold=False),
        ]
        for subj_id, _ in ordered_subjects:
            ss = sr['subj_scores'].get(subj_id, {'pct': None, 'grade': '-'})
            pct = ss['pct']
            g   = ss['grade']
            row_cells.append(_td(f"{pct:.0f}" if pct is not None else '-'))
            row_cells.append(_td(g, bold=True,
                                  color=GRADE_COLORS.get(g, C_DARK) if g != '-' else C_MGRAY))
        g_overall = sr['grade']
        row_cells += [
            _td(f"{sr['average']:.1f}", bold=True),
            _td(g_overall, bold=True, color=GRADE_COLORS.get(g_overall, C_DARK)),
            _td(sr['division']),
            _td(sr['position']),
        ]
        rows.append(row_cells)

    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    style = _ts_base(len(rows))
    # Name col left-align
    style.add('ALIGN',       (2, 0), (2, -1), 'LEFT')
    style.add('LEFTPADDING', (2, 0), (2, -1), 3)
    # Grade sub-cols: slightly tinted header
    for j, _ in enumerate(ordered_subjects):
        col_pct = 3 + j * 2
        col_grd = col_pct + 1
        style.add('BACKGROUND', (col_grd, 0), (col_grd, 0), C_BLUE2)
    tbl.setStyle(style)
    return tbl


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF — PAGE 2: SUMMARY SECTION
# ═══════════════════════════════════════════════════════════════════════════════

def _build_subject_summary_table(data, avail_w):
    col_ws = [3.5*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm,
              1.5*cm, 2.0*cm, 1.4*cm, 1.8*cm, 3.5*cm]
    used = sum(col_ws)
    if used < avail_w:
        col_ws[0] += avail_w - used

    hdr = [_th('SUBJECT'), _th('A'), _th('B'), _th('C'), _th('D'), _th('F'),
           _th('SEATS'), _th('AVERAGE'), _th('GRADE'), _th('G.P.A'), _th('COMPETENCY')]
    rows = [hdr]
    for r in data['subject_rows']:
        g = r['grade']
        rows.append([
            _td(r['subject'], align=TA_LEFT, bold=True),
            _td(r['A']), _td(r['B']), _td(r['C']), _td(r['D']), _td(r['F']),
            _td(r['seats']),
            _td(f"{r['average']:.2f}"),
            _td(g, bold=True, color=GRADE_COLORS.get(g, C_DARK)),
            _td(f"{r['gpa']:.4f}"),
            _td(r['competency'], bold=True),
        ])
    tbl = Table(rows, colWidths=col_ws, repeatRows=1)
    style = _ts_base(len(rows))
    style.add('ALIGN',       (0, 1), (0, -1), 'LEFT')
    style.add('LEFTPADDING', (0, 0), (0, -1), 4)
    tbl.setStyle(style)
    return tbl


def _build_division_block(data, avail_w):
    div_data  = data['division_summary']
    classwise = data['classwise']

    # Division table (7 cols)
    div_col_ws = [1.8*cm] + [1.5*cm]*6
    div_hdr = [_th('SEX'), _th('I'), _th('II'), _th('III'), _th('IV'), _th('0'), _th('INC')]
    div_rows = [div_hdr]
    for sex in ('M', 'F', 'TOTAL'):
        dr = div_data.get(sex, {})
        div_rows.append([
            _td(sex, bold=(sex == 'TOTAL')),
            _td(dr.get('I',0)),   _td(dr.get('II',0)),
            _td(dr.get('III',0)), _td(dr.get('IV',0)),
            _td(dr.get('0',0)),   _td(dr.get('INC',0)),
        ])
    div_tbl = Table(div_rows, colWidths=div_col_ws)
    ds = _ts_base(4)
    ds.add('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold')
    ds.add('BACKGROUND', (0, -1), (-1, -1), C_HEADER)
    div_tbl.setStyle(ds)

    # Classwise block
    cw_w = avail_w - sum(div_col_ws) - 0.5*cm
    cw_rows = [
        [_th('CLASSWISE COMPETENCY (BASED ON G.P.A)', size=8)],
        [_td(f"Overall Subjects Average: {classwise['overall_average']}, Grade: {classwise['grade']}",
             size=7.5, align=TA_LEFT)],
        [_td(f"CLASS G.P.A: {classwise['gpa']:.4f}  ({classwise['competency']})",
             size=7.5, bold=True, align=TA_LEFT)],
        [_td(f"Total Students: {data['total_students']}",
             size=7.5, align=TA_LEFT)],
    ]
    cw_tbl = Table(cw_rows, colWidths=[cw_w])
    cw_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), C_BLUE),
        ('TEXTCOLOR',     (0,0), (-1,0), WHITE),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND',    (0,1), (-1,-1), C_HEADER),
        ('GRID',          (0,0), (-1,-1), 0.4, C_MGRAY),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('ALIGN',         (0,1), (-1,-1), 'LEFT'),
        ('LEFTPADDING',   (0,1), (-1,-1), 6),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))

    gap = Table([['']], colWidths=[0.5*cm])
    gap.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0,WHITE)]))

    combined = Table([[div_tbl, gap, cw_tbl]],
                     colWidths=[sum(div_col_ws), 0.5*cm, cw_w])
    combined.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP')]))
    return combined


def _build_best_worst(data, avail_w):
    def _student_tbl(rows_data):
        hdr = [_th('SN'), _th('STUDENT NAME', align=TA_LEFT),
               _th('AVG'), _th('GR'), _th('POS')]
        half = (avail_w - 0.5*cm) / 2
        cws  = [0.7*cm, half - 4.9*cm, 1.8*cm, 1.4*cm, 1.6*cm]
        rows = [hdr]
        for r in rows_data:
            g = r['grade']
            rows.append([
                _td(r['sn']),
                _td(r['name'], align=TA_LEFT),
                _td(f"{r['average']:.1f}"),
                _td(g, bold=True, color=GRADE_COLORS.get(g, C_DARK)),
                _td(r['position']),
            ])
        t = Table(rows, colWidths=cws)
        s = _ts_base(len(rows))
        s.add('ALIGN',       (1,1), (1,-1), 'LEFT')
        s.add('LEFTPADDING', (1,0), (1,-1), 3)
        t.setStyle(s)
        return t

    half_w = (avail_w - 0.5*cm) / 2
    gap    = Table([['']], colWidths=[0.5*cm])
    gap.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0,WHITE)]))

    best_lbl  = Paragraph('LIST OF 10 BEST STUDENTS',
                           _ps('s', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE))
    worst_lbl = Paragraph('LIST OF 10 WORST STUDENTS',
                           _ps('s', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE))
    best_t  = _student_tbl(data['best_students'])
    worst_t = _student_tbl(data['worst_students'])

    outer = Table(
        [[best_lbl, '', worst_lbl],
         [best_t,   gap, worst_t]],
        colWidths=[half_w, 0.5*cm, half_w],
    )
    outer.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('SPAN',   (1,0), (1,1)),
    ]))
    return outer


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF MAIN GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_analytics_report_pdf(data: dict, school_name: str = 'MathPlatform') -> bytes:
    buf = io.BytesIO()
    MARGIN = 1.2 * cm
    PAGE_W, PAGE_H = landscape(A4)
    avail_w = PAGE_W - 2 * MARGIN

    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title='All-Subjects Analytics Report',
    )

    cls   = data['classroom']
    story = []

    def _page_header(subtitle_extra=''):
        title_str = school_name.upper()
        sub_str   = (
            f"ALL-SUBJECTS ANALYTICS REPORT  ·  {cls['name']}  ·  "
            f"{cls['academic_year']}  ·  "
            f"{str(cls['term']).replace('_',' ').title()}"
            + (f"  ·  {subtitle_extra}" if subtitle_extra else '')
        )
        h1 = Table([[_td(title_str, size=12, bold=True, color=WHITE)]],
                   colWidths=[avail_w])
        h1.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), C_BLUE),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
            ('TOPPADDING',    (0,0),(-1,-1), 6),
            ('BOTTOMPADDING', (0,0),(-1,-1), 6),
        ]))
        h2 = Table([[_td(sub_str, size=8, color=WHITE)]],
                   colWidths=[avail_w])
        h2.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), C_BLUE2),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
            ('TOPPADDING',    (0,0),(-1,-1), 3),
            ('BOTTOMPADDING', (0,0),(-1,-1), 3),
        ]))
        return [h1, h2, Spacer(1, 0.25*cm)]

    # ── PAGE 1: Student Marks Table ─────────────────────────────────────────
    story += _page_header('STUDENT MARKS')
    story.append(Paragraph('STUDENT RESULTS TABLE',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.15*cm))
    story.append(_build_marks_table(data, avail_w))

    # ── PAGE 2: Summary ─────────────────────────────────────────────────────
    story.append(PageBreak())
    story += _page_header('SUMMARY')

    story.append(Paragraph('SUBJECT SUMMARY',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.15*cm))
    story.append(_build_subject_summary_table(data, avail_w))
    story.append(Spacer(1, 0.35*cm))

    story.append(Paragraph('DIVISION SUMMARY',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.15*cm))
    story.append(_build_division_block(data, avail_w))
    story.append(Spacer(1, 0.35*cm))

    story.append(_build_best_worst(data, avail_w))

    # ── Footer ──────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.25*cm))
    story.append(HRFlowable(width='100%', thickness=0.4, color=C_MGRAY))
    story.append(Paragraph(
        f'Generated by MathPlatform  ·  {date.today().strftime("%d %B %Y")}  ·  {school_name}',
        _ps('ft', fontSize=6.5, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER),
    ))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def _xl_fill(hex6):
    return PatternFill('solid', fgColor=hex6)

def _xl_font(bold=False, color='FF000000', size=9):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def _xl_align(h='center', v='center', wrap=False, rotate=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, text_rotation=rotate)

def _xl_border(style='thin', color='FFD1D5DB'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_set(ws, row, col, value, bold=False, bg=None, fg='FF000000',
            size=9, align_h='center', wrap=False, number_format=None, rotate=0):
    c = ws.cell(row, col, value)
    c.font      = _xl_font(bold=bold, color=fg, size=size)
    c.alignment = _xl_align(h=align_h, wrap=wrap, rotate=rotate)
    c.border    = _xl_border()
    if bg:
        c.fill = _xl_fill(bg)
    if number_format:
        c.number_format = number_format
    return c

GRADE_BG = {'A':'FF10B981','B':'FF2563EB','C':'FFF59E0B','D':'FFF97316','F':'FFF43F5E'}
HDR1='1E3A5F'; HDR2='2D4F7C'; HDR3='E8F0FE'
GRAY='F3F4F6'; WHITE_XL='FFFFFF'; BLACK_XL='FF000000'


def generate_analytics_report_excel(data: dict, school_name: str = 'MathPlatform') -> bytes:
    wb = Workbook()

    # ── Sheet 1: Student Marks ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Student Marks'
    cls = data['classroom']
    ordered_subjects = data['ordered_subjects']
    n_subj = len(ordered_subjects)

    row = 1
    # Title rows
    ncols_marks = 3 + n_subj * 2 + 4
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_marks)
    _xl_set(ws1, row, 1, f'MathPlatform  ·  {school_name}',
            bold=True, bg=HDR1, fg='FFFFFFFF', size=12)
    ws1.row_dimensions[row].height = 20; row += 1

    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_marks)
    sub = (f"STUDENT RESULTS TABLE  ·  {cls['name']}  ·  "
           f"{cls['academic_year']}  ·  {str(cls['term']).replace('_',' ').upper()}")
    _xl_set(ws1, row, 1, sub, bold=True, bg=HDR2, fg='FFFFFFFF', size=10)
    ws1.row_dimensions[row].height = 16; row += 1
    row += 1  # blank

    # Column headers row
    _xl_set(ws1, row, 1, 'SN',           bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    _xl_set(ws1, row, 2, 'REG NO',       bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    _xl_set(ws1, row, 3, 'STUDENT NAME', bold=True, bg=HDR1, fg='FFFFFFFF', size=8, align_h='left')
    col_i = 4
    for _, smeta in ordered_subjects:
        label = smeta['code'] or smeta['name'][:8]
        _xl_set(ws1, row, col_i,     label, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
        _xl_set(ws1, row, col_i + 1, 'GR',  bold=True, bg=HDR2, fg='FFFFFFFF', size=8)
        col_i += 2
    _xl_set(ws1, row, col_i,     'AVG',  bold=True, bg=HDR1, fg='FFFFFFFF', size=8); col_i += 1
    _xl_set(ws1, row, col_i,     'GR',   bold=True, bg=HDR1, fg='FFFFFFFF', size=8); col_i += 1
    _xl_set(ws1, row, col_i,     'DIV',  bold=True, bg=HDR1, fg='FFFFFFFF', size=8); col_i += 1
    _xl_set(ws1, row, col_i,     'POS',  bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    ws1.row_dimensions[row].height = 28; row += 1

    # Data rows
    for i, sr in enumerate(data['student_rows']):
        bg = GRAY if i % 2 == 0 else WHITE_XL
        _xl_set(ws1, row, 1, i+1,          bg=bg, size=8)
        _xl_set(ws1, row, 2, sr['index'] or '-', bg=bg, size=8)
        _xl_set(ws1, row, 3, sr['name'],   bg=bg, size=8, align_h='left')
        col_i = 4
        for subj_id, _ in ordered_subjects:
            ss  = sr['subj_scores'].get(subj_id, {'pct': None, 'grade': '-'})
            pct = ss['pct']
            g   = ss['grade']
            pct_val = round(pct, 0) if pct is not None else ''
            _xl_set(ws1, row, col_i,     pct_val, bg=bg, size=8,
                    number_format='0' if pct_val != '' else None)
            gbg = GRADE_BG.get(g, WHITE_XL) if g != '-' else bg
            _xl_set(ws1, row, col_i + 1, g, bg=gbg, fg='FFFFFFFF' if g != '-' else BLACK_XL,
                    bold=(g != '-'), size=8)
            col_i += 2
        g_overall = sr['grade']
        _xl_set(ws1, row, col_i,     sr['average'], bg=bg, bold=True, size=8,
                number_format='0.0'); col_i += 1
        _xl_set(ws1, row, col_i,     g_overall,
                bg=GRADE_BG.get(g_overall, WHITE_XL), fg='FFFFFFFF',
                bold=True, size=8); col_i += 1
        _xl_set(ws1, row, col_i,     sr['division'],  bg=bg, size=8); col_i += 1
        _xl_set(ws1, row, col_i,     sr['position'],  bg=bg, size=8)
        ws1.row_dimensions[row].height = 14; row += 1

    # Column widths for sheet 1
    ws1.column_dimensions['A'].width = 4
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 24
    col_letter_i = 4
    for _ in ordered_subjects:
        ws1.column_dimensions[get_column_letter(col_letter_i)].width     = 6
        ws1.column_dimensions[get_column_letter(col_letter_i+1)].width   = 4
        col_letter_i += 2
    for w in [7, 4, 5, 5]:
        ws1.column_dimensions[get_column_letter(col_letter_i)].width = w
        col_letter_i += 1
    ws1.freeze_panes = 'D5'

    # ── Sheet 2: Subject Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    NCOLS = 11
    row = 1

    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 1, f'MathPlatform  ·  {school_name}',
            bold=True, bg=HDR1, fg='FFFFFFFF', size=12)
    ws2.row_dimensions[row].height = 20; row += 1

    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    sub2 = (f"ANALYTICS SUMMARY  ·  {cls['name']}  ·  "
            f"{cls['academic_year']}  ·  {str(cls['term']).replace('_',' ').upper()}")
    _xl_set(ws2, row, 1, sub2, bold=True, bg=HDR2, fg='FFFFFFFF', size=10)
    ws2.row_dimensions[row].height = 16; row += 1
    row += 1

    # Subject summary headers
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 1, 'SUBJECT SUMMARY', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.row_dimensions[row].height = 15; row += 1

    for ci, h in enumerate(['SUBJECT','A','B','C','D','F','SEATS','AVERAGE','GRADE','G.P.A','COMPETENCY'], 1):
        _xl_set(ws2, row, ci, h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    ws2.row_dimensions[row].height = 16; row += 1

    for i, r in enumerate(data['subject_rows']):
        bg = GRAY if i % 2 == 0 else WHITE_XL
        g  = r['grade']
        _xl_set(ws2, row, 1,  r['subject'],            bg=bg, size=8, align_h='left', bold=True)
        _xl_set(ws2, row, 2,  r['A'],                  bg=bg, size=8)
        _xl_set(ws2, row, 3,  r['B'],                  bg=bg, size=8)
        _xl_set(ws2, row, 4,  r['C'],                  bg=bg, size=8)
        _xl_set(ws2, row, 5,  r['D'],                  bg=bg, size=8)
        _xl_set(ws2, row, 6,  r['F'],                  bg=bg, size=8)
        _xl_set(ws2, row, 7,  r['seats'],              bg=bg, size=8)
        _xl_set(ws2, row, 8,  round(r['average'],2),   bg=bg, size=8, number_format='0.00')
        _xl_set(ws2, row, 9,  g, bg=GRADE_BG.get(g,WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
        _xl_set(ws2, row, 10, round(r['gpa'],4),       bg=bg, size=8, number_format='0.0000')
        _xl_set(ws2, row, 11, r['competency'],         bg=bg, size=8, bold=True)
        ws2.row_dimensions[row].height = 14; row += 1

    row += 1
    # Division summary
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    _xl_set(ws2, row, 1, 'DIVISION SUMMARY', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 8, 'CLASSWISE COMPETENCY', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.row_dimensions[row].height = 15; row += 1

    for ci, h in enumerate(['SEX','I','II','III','IV','0','INC'], 1):
        _xl_set(ws2, row, ci, h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    cw = data['classwise']
    ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 8,
            f"Overall Subjects Average: {cw['overall_average']}, Grade: {cw['grade']}",
            bg=HDR3, fg='FF1E3A5F', bold=True, size=8, align_h='left')
    ws2.row_dimensions[row].height = 15; row += 1

    div_summary = data['division_summary']
    cw_extra = [
        f"CLASS G.P.A: {cw['gpa']:.4f}  ({cw['competency']})",
        f"Total Students: {data['total_students']}",
        '',
    ]
    for i, sex in enumerate(('M', 'F', 'TOTAL')):
        bg_row = 'D1E8FF' if sex == 'TOTAL' else (GRAY if i % 2 == 0 else WHITE_XL)
        bd     = sex == 'TOTAL'
        dr     = div_summary.get(sex, {})
        _xl_set(ws2, row, 1, sex,             bg=bg_row, bold=bd, size=8)
        for ci2, div in enumerate(['I','II','III','IV','0','INC'], 2):
            _xl_set(ws2, row, ci2, dr.get(div, 0), bg=bg_row, bold=bd, size=8)
        ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
        _xl_set(ws2, row, 8, cw_extra[i],
                bg=HDR3, fg='FF1E3A5F', bold=(i==0), size=8, align_h='left')
        ws2.row_dimensions[row].height = 14; row += 1

    row += 1
    # Best / Worst
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _xl_set(ws2, row, 1, 'LIST OF 10 BEST STUDENTS', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.merge_cells(start_row=row, start_column=7, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 7, 'LIST OF 10 WORST STUDENTS', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.cell(row, 6).fill = _xl_fill(WHITE_XL)
    ws2.row_dimensions[row].height = 15; row += 1

    sw_hdrs = ['SN','STUDENT NAME','AVERAGE','GRADE','POSITION']
    for ci, h in enumerate(sw_hdrs, 1):
        _xl_set(ws2, row, ci,     h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
        _xl_set(ws2, row, ci + 6, h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    ws2.cell(row, 6).fill = _xl_fill(WHITE_XL)
    ws2.row_dimensions[row].height = 14; row += 1

    max_rows = max(len(data['best_students']), len(data['worst_students']))
    for i in range(max_rows):
        bg = GRAY if i % 2 == 0 else WHITE_XL
        if i < len(data['best_students']):
            br = data['best_students'][i]
            g  = br['grade']
            _xl_set(ws2, row, 1, br['sn'],       bg=bg, size=8)
            _xl_set(ws2, row, 2, br['name'],      bg=bg, size=8, align_h='left')
            _xl_set(ws2, row, 3, br['average'],   bg=bg, size=8, number_format='0.0')
            _xl_set(ws2, row, 4, g, bg=GRADE_BG.get(g,WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
            _xl_set(ws2, row, 5, br['position'],  bg=bg, size=8)
        if i < len(data['worst_students']):
            wr = data['worst_students'][i]
            g  = wr['grade']
            _xl_set(ws2, row, 7,  wr['sn'],       bg=bg, size=8)
            _xl_set(ws2, row, 8,  wr['name'],     bg=bg, size=8, align_h='left')
            _xl_set(ws2, row, 9,  wr['average'],  bg=bg, size=8, number_format='0.0')
            _xl_set(ws2, row, 10, g, bg=GRADE_BG.get(g,WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
            _xl_set(ws2, row, 11, wr['position'], bg=bg, size=8)
        ws2.cell(row, 6).fill = _xl_fill(WHITE_XL)
        ws2.row_dimensions[row].height = 14; row += 1

    # Footer
    row += 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 1,
            f'Generated by MathPlatform  ·  {date.today().strftime("%d %B %Y")}  ·  {school_name}',
            bg=HDR1, fg='FFFFFFFF', size=7)
    ws2.row_dimensions[row].height = 13

    # Column widths sheet 2
    col_w2 = [22, 4.5, 4.5, 4.5, 4.5, 4.5, 6, 10, 7, 9, 16]
    for ci, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

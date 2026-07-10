"""
Analytics Report Engine — All-Subjects Summary Report
Generates NECTA-style subject summary tables for PDF and Excel.

REPORT STRUCTURE (2 sections):
  PAGE 1 — STUDENT MARKS TABLE
    SN | REG NO | STUDENT NAME | [SUBJ score/grade] x N | TOTAL | AVERAGE | GRADE | DIV | POS
  PAGE 2 — SUMMARY
    Subject Summary table
    Division Summary  +  Classwise Competency
    List of 10 Best   +  List of 10 Worst
"""
import io
from datetime import date
from collections import defaultdict

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, String, Rect
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED GRADE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _letter_grade(pct):
    if pct is None: return '-'
    if pct >= 75:   return 'A'
    if pct >= 65:   return 'B'
    if pct >= 45:   return 'C'
    if pct >= 30:   return 'D'
    return 'F'

def _gpa(pct):
    if pct is None: return None
    if pct >= 75:   return 1.0
    if pct >= 65:   return 2.0
    if pct >= 45:   return 3.0
    if pct >= 30:   return 4.0
    return 5.0

def _competency(gpa):
    if gpa is None:  return '-'
    if gpa <= 1.5:   return 'EXCELLENT'
    if gpa <= 2.0:   return 'VERY GOOD'
    if gpa <= 2.5:   return 'GOOD'
    if gpa <= 3.5:   return 'SATISFACTORY'
    if gpa <= 4.0:   return 'AVERAGE'
    return 'FAIL'

MIN_DIVISION_SUBJECTS = 7  # NECTA divisions are only awarded on 7+ subjects


def _division_from_points(total_points):
    """NECTA-style division from the summed GPA points (1=A best ... 5=F
    worst per subject) of a student's best 7 subjects. Returns '-' when
    total_points is None (fewer than MIN_DIVISION_SUBJECTS subjects taken —
    a division cannot be honestly computed, so this is left blank rather
    than defaulting to '0'/fail, which would misreport an incomplete record
    as an outright failure)."""
    if total_points is None:
        return '-'
    if total_points <= 17: return 'I'
    if total_points <= 21: return 'II'
    if total_points <= 25: return 'III'
    if total_points <= 33: return 'IV'
    return '0'


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_analytics_report_data(classroom_id, academic_year=None, term=None,
                                 subject_id=None, created_by_id=None, exam_id=None):
    from django.db.models import FloatField, ExpressionWrapper, F, Q
    from mathapi.apps.students.models import StudentProfile, Classroom
    from mathapi.apps.exams.models import Exam, ExamScore

    classroom = Classroom.objects.select_related('grade_level').get(id=classroom_id)

    pct_expr = ExpressionWrapper(
        F('score') * 100.0 / F('exam__max_score'),
        output_field=FloatField(),
    )

    # ── Scope filter ────────────────────────────────────────────────────────
    base_filter = Q(
        student__classroom_id=classroom_id,
        is_absent=False,
        exam__max_score__gt=0,
        exam__is_deleted=False,
    )
    if exam_id:
        base_filter &= Q(exam_id=exam_id)
    else:
        if academic_year:
            base_filter &= Q(exam__academic_year=academic_year)
        if term:
            base_filter &= Q(exam__term=term)
        if created_by_id:
            base_filter &= Q(exam__created_by_id=created_by_id)
        if subject_id:
            base_filter &= Q(exam__subject_id=subject_id)

    # ── Pull all scores ─────────────────────────────────────────────────────
    scores_qs = list(
        ExamScore.objects
        .filter(base_filter)
        .annotate(pct=pct_expr)
        .values(
            'student_id',
            'student__user__first_name',
            'student__user__last_name',
            'student__index_number',
            'exam__subject_id',
            'exam__subject__name',
            'exam__subject__code',
            'score',
            'exam__max_score',
            'pct',
        )
        .order_by('student__user__last_name', 'student__user__first_name')
    )

    # ── Aggregate per (student, subject) ────────────────────────────────────
    # Store lists of (score, max_score, pct) per (student_id, subject_id)
    student_subject_raw: dict = defaultdict(list)   # → [(score, max_score, pct), ...]
    student_meta: dict = {}
    subject_meta: dict = {}

    for row in scores_qs:
        sid  = row['student_id']
        subj = row['exam__subject_id']
        pct  = row['pct']
        if pct is None:
            continue
        student_subject_raw[(sid, subj)].append((
            float(row['score']),
            float(row['exam__max_score']),
            float(pct),
        ))
        if sid not in student_meta:
            fn = row['student__user__first_name'] or ''
            ln = row['student__user__last_name'] or ''
            student_meta[sid] = {
                'name': f'{fn} {ln}'.strip() or f'Student {sid}',
                'index': row['student__index_number'] or '',
            }
        if subj and subj not in subject_meta:
            subject_meta[subj] = {
                'name': row['exam__subject__name'] or f'Subject {subj}',
                'code': row['exam__subject__code'] or '',
            }

    if not student_meta:
        raise ValueError(
            'No scored exams found for this classroom with the selected filters. '
            'Ensure exams have been entered with marks for students in this classroom.'
        )

    # avg pct per (student, subject)
    avg_ss: dict = {}
    for (sid, subj), entries in student_subject_raw.items():
        total_score = sum(e[0] for e in entries)
        total_max   = sum(e[1] for e in entries)
        avg_pct     = (total_score / total_max * 100) if total_max else 0.0
        avg_ss[(sid, subj)] = round(avg_pct, 1)

    # ordered subject list (by code/name)
    ordered_subjects = sorted(
        subject_meta.items(),
        key=lambda x: x[1]['code'] or x[1]['name']
    )  # [(subj_id, {name, code}), ...]

    # ── Per-student row data (for marks table) ──────────────────────────────
    student_rows = []
    for sid, smeta in sorted(student_meta.items(),
                              key=lambda x: x[1]['name']):
        subj_scores = {}  # subj_id → {pct, grade}
        for subj_id, _ in ordered_subjects:
            pct = avg_ss.get((sid, subj_id))
            subj_scores[subj_id] = {
                'pct':   pct,
                'grade': _letter_grade(pct) if pct is not None else '-',
            }
        # collect all pcts for this student
        student_pcts = [avg_ss[(sid, subj_id)]
                        for subj_id, _ in ordered_subjects
                        if (sid, subj_id) in avg_ss]
        if not student_pcts:
            continue
        avg_pct = round(sum(student_pcts) / len(student_pcts), 1)
        avg_gpa = round(sum(_gpa(p) for p in student_pcts) / len(student_pcts), 4)

        # ── Division: best 7 subjects only ──────────────────────────────
        # Sort this student's per-subject GPA points ascending (1=A is the
        # best point score, 5=F the worst) and sum the lowest 7 — i.e. their
        # 7 best-performed subjects — to get the NECTA-style point total.
        # If they have fewer than 7 subjects recorded, a division genuinely
        # can't be awarded, so total_points stays None → division shows '-'.
        subject_points = sorted(_gpa(p) for p in student_pcts)
        if len(subject_points) >= MIN_DIVISION_SUBJECTS:
            total_points = sum(subject_points[:MIN_DIVISION_SUBJECTS])
        else:
            total_points = None

        student_rows.append({
            'student_id':   sid,
            'name':         smeta['name'],
            'index':        smeta['index'],
            'subj_scores':  subj_scores,        # subj_id → {pct, grade}
            'total_pct':    round(sum(student_pcts), 1),
            'average':      avg_pct,
            'grade':        _letter_grade(avg_pct),
            'gpa':          avg_gpa,
            'division_points': total_points,
            'division':     _division_from_points(total_points),
        })

    # Assign positions (rank by average descending)
    student_rows.sort(key=lambda r: -r['average'])
    for i, r in enumerate(student_rows):
        r['position'] = i + 1
    # Re-sort alphabetically for display
    student_rows.sort(key=lambda r: r['name'])

    # ── Subject summary ─────────────────────────────────────────────────────
    subject_rows = []
    for subj_id, smeta in ordered_subjects:
        pcts = [avg_ss[(sid, subj_id)]
                for sid in student_meta
                if (sid, subj_id) in avg_ss]
        if not pcts:
            continue
        grades = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        for p in pcts:
            g = _letter_grade(p)
            if g in grades:
                grades[g] += 1
        avg      = round(sum(pcts) / len(pcts), 2)
        gpa_val  = round(sum(_gpa(p) for p in pcts) / len(pcts), 4)
        subject_rows.append({
            'subject':    smeta['code'] or smeta['name'],
            'full_name':  smeta['name'],
            'A': grades['A'], 'B': grades['B'], 'C': grades['C'],
            'D': grades['D'], 'F': grades['F'],
            'seats':      len(pcts),
            'average':    avg,
            'grade':      _letter_grade(avg),
            'gpa':        round(gpa_val, 4),
            'competency': _competency(gpa_val),
        })

    # ── Per-student overall ─────────────────────────────────────────────────
    student_overall = {r['student_id']: r for r in student_rows}

    # ── Class-wide overall grade distribution (one entry per student, based
    #    on their overall average — used by the grade-distribution chart) ───
    grade_distribution = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    for r in student_rows:
        if r['grade'] in grade_distribution:
            grade_distribution[r['grade']] += 1

    # ── NECTA-format grade distribution ─────────────────────────────────────
    # NECTA never grades a student's "overall average" — every grade is
    # awarded per subject entry. Its published Grade Distribution is the sum
    # of A/B/C/D/F across every subject *entry* sat by the class (the same
    # counts already sitting in each subject_row, just totalled). That's
    # what real CSEE/FTNA analysis sheets show, so the report's grade
    # distribution table/chart is sourced from here rather than from the
    # per-student-average figure above.
    necta_grade_distribution = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    for sub in subject_rows:
        for g in necta_grade_distribution:
            necta_grade_distribution[g] += sub[g]

    # ── Division summary ────────────────────────────────────────────────────
    # NOTE: StudentProfile has no gender field, so a real M/F split cannot be
    # computed. Previously this hardcoded M to 0 and put the *entire* class
    # total into F, which silently mislabeled every student as female. Until
    # gender is tracked on StudentProfile, only report the honest TOTAL.
    div_map: dict = defaultdict(int)
    for r in student_rows:
        div_map[r['division']] += 1

    division_summary = {
        'TOTAL': {div: div_map.get(div, 0) for div in ('I', 'II', 'III', 'IV', '0')},
        # Students with fewer than 7 subjects recorded — no division can be
        # awarded for them (shown as '-' on their row), so they're broken
        # out here rather than silently missing from the division counts.
        'INCOMPLETE': div_map.get('-', 0),
    }

    # ── Classwise competency ────────────────────────────────────────────────
    all_avgs = [r['average'] for r in student_rows]
    all_gpas = [r['gpa'] for r in student_rows]
    class_avg = round(sum(all_avgs) / len(all_avgs), 1) if all_avgs else 0.0
    class_gpa = round(sum(all_gpas) / len(all_gpas), 4) if all_gpas else 0.0

    classwise = {
        'overall_average': class_avg,
        'grade':           _letter_grade(class_avg),
        'gpa':             class_gpa,
        'competency':      _competency(class_gpa),
    }

    # ── Best / worst 10 ────────────────────────────────────────────────────
    # Cap each list so they never overlap on small classes: previously
    # ranked[:10] and ranked[-10:] shared students whenever a classroom had
    # <=20 students (very common), making "Best 10" and "Worst 10" show the
    # same names.
    ranked = sorted(student_rows, key=lambda r: -r['average'])
    half = max(len(ranked) // 2, 0)
    best_n = min(10, half) if len(ranked) < 20 else 10
    worst_n = min(10, len(ranked) - best_n)
    best_10 = [
        {'sn': i+1, 'name': r['name'], 'average': r['average'],
         'grade': r['grade'], 'position': r['position']}
        for i, r in enumerate(ranked[:best_n])
    ]
    worst_10 = [
        {'sn': i+1, 'name': r['name'], 'average': r['average'],
         'grade': r['grade'], 'position': r['position']}
        for i, r in enumerate(reversed(ranked[len(ranked)-worst_n:])) if worst_n > 0
    ]

    return {
        'classroom': {
            'name':          classroom.name,
            'grade':         classroom.grade_level.name if classroom.grade_level_id else '',
            'academic_year': academic_year or classroom.academic_year,
            'term':          term or 'All Terms',
        },
        'ordered_subjects': ordered_subjects,    # [(subj_id, {name, code}), ...]
        'student_rows':     student_rows,         # full per-student data
        'subject_rows':     subject_rows,         # aggregated subject summary
        'division_summary': division_summary,
        'grade_distribution': grade_distribution,
        'necta_grade_distribution': necta_grade_distribution,
        'classwise':        classwise,
        'best_students':    best_10,
        'worst_students':   worst_10,
        'total_students':   len(student_rows),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF COLOURS & HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

C_DARK   = colors.HexColor('#0a0a0f')
C_BLUE   = colors.HexColor('#1e3a5f')
C_BLUE2  = colors.HexColor('#2d4f7c')
C_LGRAY  = colors.HexColor('#f3f4f6')
C_MGRAY  = colors.HexColor('#d1d5db')
C_GREEN  = colors.HexColor('#10b981')
C_AMBER  = colors.HexColor('#f59e0b')
C_RED    = colors.HexColor('#f43f5e')
C_HEADER = colors.HexColor('#e8f0fe')
WHITE    = colors.white
BLACK    = colors.black

GRADE_COLORS = {
    'A': C_GREEN,
    'B': colors.HexColor('#2563eb'),
    'C': C_AMBER,
    'D': colors.HexColor('#f97316'),
    'F': C_RED,
}

def _ps(name, **kw):
    d = dict(fontName='Helvetica', fontSize=8, textColor=C_DARK,
             spaceAfter=0, spaceBefore=0, leading=10)
    d.update(kw)
    return ParagraphStyle(name, **d)

def _th(text, size=7.5, bold=True, align=TA_CENTER, color=WHITE):
    return Paragraph(text,
        _ps('th', fontName='Helvetica-Bold' if bold else 'Helvetica',
            fontSize=size, alignment=align, textColor=color))

def _td(text, size=7.5, bold=False, align=TA_CENTER, color=C_DARK):
    return Paragraph(str(text) if text is not None else '',
        _ps('td', fontName='Helvetica-Bold' if bold else 'Helvetica',
            fontSize=size, alignment=align, textColor=color))

def _ts_base(nrows):
    return TableStyle([
        ('FONTNAME',       (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE',       (0, 0), (-1, -1), 7.5),
        ('GRID',           (0, 0), (-1, -1), 0.4, C_MGRAY),
        ('BACKGROUND',     (0, 0), (-1, 0),  C_BLUE),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  WHITE),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, nrows-1), [C_LGRAY, WHITE]),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',     (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 2),
    ])


def _header_block(school_name, classroom_info, subtitle=''):
    """Returns (header_table, subtitle_table)."""
    cls = classroom_info
    title = school_name.upper()
    sub   = subtitle or (
        f"ALL-SUBJECTS ANALYTICS REPORT  ·  {cls['name']}  ·  "
        f"{cls['academic_year']}  ·  {str(cls['term']).replace('_',' ').upper()}"
    )
    return title, sub


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF — PAGE 1: STUDENT MARKS TABLE
# ═══════════════════════════════════════════════════════════════════════════════

def _build_marks_table(data, avail_w):
    """Build the student-by-subject marks table flowable."""
    ordered_subjects = data['ordered_subjects']   # [(id, {name, code}), ...]
    student_rows     = data['student_rows']

    n_subj = len(ordered_subjects)

    # ── Column widths ────────────────────────────────────────────────────────
    # Fixed cols: SN(0.5) + REG(2.0) + NAME(4.0) + ... + AVG(1.4) + GRD(1.0) + DIV(0.7) + POS(0.8)
    W_SN   = 0.55 * cm
    W_REG  = 2.00 * cm
    W_NAME = 4.00 * cm
    W_AVG  = 1.40 * cm
    W_GRD  = 0.90 * cm
    W_DIV  = 0.75 * cm
    W_POS  = 0.75 * cm
    fixed  = W_SN + W_REG + W_NAME + W_AVG + W_GRD + W_DIV + W_POS

    # Each subject gets two sub-columns: score% + grade.
    # IMPORTANT: no minimum-width floor here. A previous version enforced
    # max(1.6cm, remaining/n_subj), which meant that once a class had more
    # than ~11 subjects the table's total width exceeded avail_w and
    # ReportLab silently clipped/overflowed the right-most subject columns
    # off the page. Dividing evenly guarantees the table always fits; font
    # size shrinks instead so many-subject classes stay legible rather than
    # losing columns entirely.
    remaining = avail_w - fixed
    W_SUBJ_PAIR = remaining / max(n_subj, 1)
    if W_SUBJ_PAIR >= 1.6 * cm:
        subj_font = 7.5
    elif W_SUBJ_PAIR >= 1.1 * cm:
        subj_font = 6.5
    elif W_SUBJ_PAIR >= 0.8 * cm:
        subj_font = 5.5
    else:
        subj_font = 4.5
    W_PCT  = W_SUBJ_PAIR * 0.60
    W_GRD2 = W_SUBJ_PAIR * 0.40

    col_widths = [W_SN, W_REG, W_NAME]
    for _ in ordered_subjects:
        col_widths += [W_PCT, W_GRD2]
    col_widths += [W_AVG, W_GRD, W_DIV, W_POS]

    # ── Header row ───────────────────────────────────────────────────────────
    hdr = [
        _th('SN'),
        _th('REG NO'),
        _th('STUDENT NAME', align=TA_LEFT),
    ]
    for _, smeta in ordered_subjects:
        # Truncate labels harder as columns get narrower so codes don't
        # wrap and blow out the row height.
        max_chars = 6 if subj_font >= 6.5 else (4 if subj_font >= 5.5 else 3)
        label = (smeta['code'] or smeta['name'])[:max_chars]
        hdr += [_th(label, size=subj_font), _th('GR', size=subj_font)]
    hdr += [_th('AVG'), _th('GR'), _th('DIV'), _th('POS')]

    rows = [hdr]

    for i, sr in enumerate(student_rows):
        bg = C_LGRAY if i % 2 == 0 else WHITE
        row_cells = [
            _td(i + 1),
            _td(sr['index'] or '-'),
            _td(sr['name'], align=TA_LEFT, bold=False),
        ]
        for subj_id, _ in ordered_subjects:
            ss = sr['subj_scores'].get(subj_id, {'pct': None, 'grade': '-'})
            pct = ss['pct']
            g   = ss['grade']
            row_cells.append(_td(f"{pct:.0f}" if pct is not None else '-', size=subj_font))
            row_cells.append(_td(g, size=subj_font, bold=True,
                                  color=GRADE_COLORS.get(g, C_DARK) if g != '-' else C_MGRAY))
        g_overall = sr['grade']
        row_cells += [
            _td(f"{sr['average']:.1f}", bold=True),
            _td(g_overall, bold=True, color=GRADE_COLORS.get(g_overall, C_DARK)),
            _td(sr['division']),
            _td(sr['position']),
        ]
        rows.append(row_cells)

    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    style = _ts_base(len(rows))
    # Name col left-align
    style.add('ALIGN',       (2, 0), (2, -1), 'LEFT')
    style.add('LEFTPADDING', (2, 0), (2, -1), 3)
    # Grade sub-cols: slightly tinted header
    for j, _ in enumerate(ordered_subjects):
        col_pct = 3 + j * 2
        col_grd = col_pct + 1
        style.add('BACKGROUND', (col_grd, 0), (col_grd, 0), C_BLUE2)
    tbl.setStyle(style)
    return tbl


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF — PAGE 2: SUMMARY SECTION
# ═══════════════════════════════════════════════════════════════════════════════

def _build_subject_summary_table(data, avail_w):
    col_ws = [3.5*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm, 1.1*cm,
              1.5*cm, 2.0*cm, 1.4*cm, 1.8*cm, 3.5*cm]
    used = sum(col_ws)
    if used < avail_w:
        col_ws[0] += avail_w - used

    hdr = [_th('SUBJECT'), _th('A'), _th('B'), _th('C'), _th('D'), _th('F'),
           _th('SEATS'), _th('AVERAGE'), _th('GRADE'), _th('G.P.A'), _th('COMPETENCY')]
    rows = [hdr]
    for r in data['subject_rows']:
        g = r['grade']
        rows.append([
            _td(r['subject'], align=TA_LEFT, bold=True),
            _td(r['A']), _td(r['B']), _td(r['C']), _td(r['D']), _td(r['F']),
            _td(r['seats']),
            _td(f"{r['average']:.2f}"),
            _td(g, bold=True, color=GRADE_COLORS.get(g, C_DARK)),
            _td(f"{r['gpa']:.4f}"),
            _td(r['competency'], bold=True),
        ])

    # NECTA-style TOTAL row: the sum of A/B/C/D/F across every subject entry
    # in the class (this is the same figure published as "Grade
    # Distribution" on real CSEE/FTNA analysis sheets).
    ngd = data['necta_grade_distribution']
    total_seats = sum(ngd.values())
    rows.append([
        _td('TOTAL', align=TA_LEFT, bold=True),
        _td(ngd['A'], bold=True), _td(ngd['B'], bold=True), _td(ngd['C'], bold=True),
        _td(ngd['D'], bold=True), _td(ngd['F'], bold=True),
        _td(total_seats, bold=True),
        _td(''), _td(''), _td(''), _td(''),
    ])

    tbl = Table(rows, colWidths=col_ws, repeatRows=1)
    style = _ts_base(len(rows))
    style.add('ALIGN',       (0, 1), (0, -1), 'LEFT')
    style.add('LEFTPADDING', (0, 0), (0, -1), 4)
    style.add('BACKGROUND',  (0, -1), (-1, -1), C_HEADER)
    style.add('FONTNAME',    (0, -1), (-1, -1), 'Helvetica-Bold')
    tbl.setStyle(style)
    return tbl


def _build_division_block(data, avail_w):
    div_data  = data['division_summary']
    classwise = data['classwise']

    # Division table (6 cols) — only TOTAL is real (no gender data tracked);
    # see build_analytics_report_data() for why M/F rows were removed.
    div_col_ws = [1.8*cm] + [1.3*cm]*6
    div_hdr = [_th('DIVISION'), _th('I'), _th('II'), _th('III'), _th('IV'), _th('0'), _th('N/A')]
    div_rows = [div_hdr]
    dr = div_data.get('TOTAL', {})
    div_rows.append([
        _td('TOTAL', bold=True),
        _td(dr.get('I',0)),   _td(dr.get('II',0)),
        _td(dr.get('III',0)), _td(dr.get('IV',0)),
        _td(dr.get('0',0)),   _td(div_data.get('INCOMPLETE', 0)),
    ])
    div_tbl = Table(div_rows, colWidths=div_col_ws)
    ds = _ts_base(len(div_rows))
    ds.add('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold')
    ds.add('BACKGROUND', (0, -1), (-1, -1), C_HEADER)
    div_tbl.setStyle(ds)

    # Classwise block
    cw_w = avail_w - sum(div_col_ws) - 0.5*cm
    cw_rows = [
        [_th('CLASSWISE COMPETENCY (BASED ON G.P.A)', size=8)],
        [_td(f"Overall Subjects Average: {classwise['overall_average']}, Grade: {classwise['grade']}",
             size=7.5, align=TA_LEFT)],
        [_td(f"CLASS G.P.A: {classwise['gpa']:.4f}  ({classwise['competency']})",
             size=7.5, bold=True, align=TA_LEFT)],
        [_td(f"Total Students: {data['total_students']}",
             size=7.5, align=TA_LEFT)],
    ]
    cw_tbl = Table(cw_rows, colWidths=[cw_w])
    cw_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), C_BLUE),
        ('TEXTCOLOR',     (0,0), (-1,0), WHITE),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND',    (0,1), (-1,-1), C_HEADER),
        ('GRID',          (0,0), (-1,-1), 0.4, C_MGRAY),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('ALIGN',         (0,1), (-1,-1), 'LEFT'),
        ('LEFTPADDING',   (0,1), (-1,-1), 6),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))

    gap = Table([['']], colWidths=[0.5*cm])
    gap.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0,WHITE)]))

    combined = Table([[div_tbl, gap, cw_tbl]],
                     colWidths=[sum(div_col_ws), 0.5*cm, cw_w])
    combined.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP')]))
    return combined


def _build_best_worst(data, avail_w):
    def _student_tbl(rows_data):
        hdr = [_th('SN'), _th('STUDENT NAME', align=TA_LEFT),
               _th('AVG'), _th('GR'), _th('POS')]
        half = (avail_w - 0.5*cm) / 2
        cws  = [0.7*cm, half - 4.9*cm, 1.8*cm, 1.4*cm, 1.6*cm]
        rows = [hdr]
        for r in rows_data:
            g = r['grade']
            rows.append([
                _td(r['sn']),
                _td(r['name'], align=TA_LEFT),
                _td(f"{r['average']:.1f}"),
                _td(g, bold=True, color=GRADE_COLORS.get(g, C_DARK)),
                _td(r['position']),
            ])
        t = Table(rows, colWidths=cws)
        s = _ts_base(len(rows))
        s.add('ALIGN',       (1,1), (1,-1), 'LEFT')
        s.add('LEFTPADDING', (1,0), (1,-1), 3)
        t.setStyle(s)
        return t

    half_w = (avail_w - 0.5*cm) / 2
    gap    = Table([['']], colWidths=[0.5*cm])
    gap.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0,WHITE)]))

    best_lbl  = Paragraph('LIST OF 10 BEST STUDENTS',
                           _ps('s', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE))
    worst_lbl = Paragraph('LIST OF 10 WORST STUDENTS',
                           _ps('s', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE))
    best_t  = _student_tbl(data['best_students'])
    worst_t = _student_tbl(data['worst_students'])

    outer = Table(
        [[best_lbl, '', worst_lbl],
         [best_t,   gap, worst_t]],
        colWidths=[half_w, 0.5*cm, half_w],
    )
    outer.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('SPAN',   (1,0), (1,1)),
    ]))
    return outer


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF CHARTS  (subject averages · grade distribution · division distribution)
# ═══════════════════════════════════════════════════════════════════════════════

def _subject_average_chart(data, width, height):
    """Vertical bar chart: average % per subject, bars coloured by the same
    grade band each subject actually earned, with value labels on each bar
    and a legend mapping colour → grade band."""
    subjects = data['subject_rows']
    d = Drawing(width, height)

    if not subjects:
        d.add(String(width/2, height/2, 'No subject data available',
                      fontSize=9, fillColor=C_MGRAY, textAnchor='middle'))
        return d

    chart_w = width - 3.6*cm
    chart = VerticalBarChart()
    chart.x = 1.5*cm
    chart.y = 1.3*cm
    chart.width = chart_w
    chart.height = height - 2.3*cm
    chart.data = [[s['average'] for s in subjects]]
    chart.categoryAxis.categoryNames = [
        (s['subject'][:10] + '…') if len(s['subject']) > 11 else s['subject']
        for s in subjects
    ]
    chart.categoryAxis.labels.fontSize = 7.5
    chart.categoryAxis.labels.fontName = 'Helvetica-Bold'
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = 100
    chart.valueAxis.valueSteps = [0, 25, 50, 75, 100]
    chart.valueAxis.labelTextFormat = '%d%%'
    chart.valueAxis.labels.fontSize = 7
    chart.barWidth = 10
    chart.groupSpacing = 8
    chart.barLabels.nudge = 8
    chart.barLabelFormat = '%0.1f%%'
    chart.barLabels.fontSize = 6.5
    chart.barLabels.fillColor = C_DARK
    for i, s in enumerate(subjects):
        chart.bars[(0, i)].fillColor = GRADE_COLORS.get(s['grade'], C_MGRAY)
    d.add(chart)

    # ── Legend / key: colour → grade band ───────────────────────────────────
    legend = Legend()
    legend.x = width - 1.9*cm
    legend.y = height - 0.6*cm
    legend.dx = 7
    legend.dy = 7
    legend.fontSize = 6.5
    legend.fontName = 'Helvetica'
    legend.alignment = 'right'
    legend.columnMaximum = 6
    legend.colorNamePairs = [
        (GRADE_COLORS['A'], 'A ≥75%'), (GRADE_COLORS['B'], 'B 65-74%'),
        (GRADE_COLORS['C'], 'C 45-64%'), (GRADE_COLORS['D'], 'D 30-44%'),
        (GRADE_COLORS['F'], 'F <30%'),
    ]
    d.add(legend)
    d.add(String(width/2, height - 0.35*cm, 'Subject-wise Class Average',
                  fontSize=8, fontName='Helvetica-Bold', fillColor=C_DARK, textAnchor='middle'))
    return d


def _grade_distribution_chart(data, width, height):
    """Pie chart of the NECTA-format grade distribution — A/B/C/D/F counted
    across every subject entry sat by the class, matching how CSEE/FTNA
    analysis sheets report it — with a legend showing each grade's colour,
    label and count."""
    dist = data['necta_grade_distribution']
    d = Drawing(width, height)
    total = sum(dist.values())
    if not total:
        d.add(String(width/2, height/2, 'No graded students yet',
                      fontSize=9, fillColor=C_MGRAY, textAnchor='middle'))
        return d

    order = ['A', 'B', 'C', 'D', 'F']
    labels = [g for g in order if dist.get(g)]
    values = [dist[g] for g in labels]

    pie = Pie()
    pie.x = 0.4*cm
    pie.y = 0.5*cm
    pie.width = height - 1.6*cm
    pie.height = height - 1.6*cm
    pie.data = values
    pie.labels = [f'{v} ({v/total*100:.0f}%)' for v in values]
    pie.simpleLabels = False
    pie.sideLabels = False
    pie.slices.fontSize = 7
    pie.slices.fontColor = C_DARK
    pie.slices.strokeColor = WHITE
    pie.slices.strokeWidth = 1
    for i, g in enumerate(labels):
        pie.slices[i].fillColor = GRADE_COLORS.get(g, C_MGRAY)
    d.add(pie)

    legend = Legend()
    legend.x = pie.width + 1.0*cm
    legend.y = height - 0.7*cm
    legend.dx = 7
    legend.dy = 7
    legend.fontSize = 7
    legend.fontName = 'Helvetica'
    legend.alignment = 'right'
    legend.columnMaximum = 5
    legend.colorNamePairs = [
        (GRADE_COLORS.get(g, C_MGRAY), f'{g}  —  {dist[g]} student{"s" if dist[g] != 1 else ""}')
        for g in labels
    ]
    d.add(legend)
    d.add(String(width/2, height - 0.2*cm, 'Grade Distribution (NECTA format)',
                  fontSize=8, fontName='Helvetica-Bold', fillColor=C_DARK, textAnchor='middle'))
    return d


DIVISION_COLORS = {
    'I':   C_GREEN,
    'II':  colors.HexColor('#2563eb'),
    'III': C_AMBER,
    'IV':  colors.HexColor('#f97316'),
    '0':   C_RED,
}


def _division_distribution_chart(data, width, height):
    """Bar chart of student counts per NECTA-style division (I best → 0
    fail), with value labels and a legend explaining each division."""
    div_row = data['division_summary'].get('TOTAL', {})
    d = Drawing(width, height)
    order = ['I', 'II', 'III', 'IV', '0']
    values = [div_row.get(k, 0) for k in order]
    if not sum(values):
        d.add(String(width/2, height/2, 'No division data available',
                      fontSize=9, fillColor=C_MGRAY, textAnchor='middle'))
        return d

    chart = VerticalBarChart()
    chart.x = 1.3*cm
    chart.y = 1.3*cm
    chart.width = width - 3.4*cm
    chart.height = height - 2.3*cm
    chart.data = [values]
    chart.categoryAxis.categoryNames = [f'Div {k}' if k != '0' else 'Div 0 (Fail)' for k in order]
    chart.categoryAxis.labels.fontSize = 7
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labels.fontSize = 7
    chart.barWidth = 14
    chart.groupSpacing = 10
    chart.barLabels.nudge = 8
    chart.barLabelFormat = '%d'
    chart.barLabels.fontSize = 7.5
    chart.barLabels.fillColor = C_DARK
    for i, k in enumerate(order):
        chart.bars[(0, i)].fillColor = DIVISION_COLORS.get(k, C_MGRAY)
    d.add(chart)

    legend = Legend()
    legend.x = width - 1.5*cm
    legend.y = height - 0.6*cm
    legend.dx = 7
    legend.dy = 7
    legend.fontSize = 6.5
    legend.fontName = 'Helvetica'
    legend.alignment = 'right'
    legend.columnMaximum = 5
    legend.colorNamePairs = [(DIVISION_COLORS[k], f'Div {k}' if k != '0' else 'Div 0 (Fail)') for k in order]
    d.add(legend)
    d.add(String(width/2, height - 0.35*cm, 'Division Distribution',
                  fontSize=8, fontName='Helvetica-Bold', fillColor=C_DARK, textAnchor='middle'))
    return d


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF MAIN GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_analytics_report_pdf(data: dict, school_name: str = 'MathPlatform') -> bytes:
    buf = io.BytesIO()
    MARGIN = 1.2 * cm
    PAGE_W, PAGE_H = landscape(A4)
    avail_w = PAGE_W - 2 * MARGIN

    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title='All-Subjects Analytics Report',
    )

    cls   = data['classroom']
    story = []

    def _page_header(subtitle_extra=''):
        title_str = school_name.upper()
        sub_str   = (
            f"ALL-SUBJECTS ANALYTICS REPORT  ·  {cls['name']}  ·  "
            f"{cls['academic_year']}  ·  "
            f"{str(cls['term']).replace('_',' ').title()}"
            + (f"  ·  {subtitle_extra}" if subtitle_extra else '')
        )
        h1 = Table([[_td(title_str, size=12, bold=True, color=WHITE)]],
                   colWidths=[avail_w])
        h1.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), C_BLUE),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
            ('TOPPADDING',    (0,0),(-1,-1), 6),
            ('BOTTOMPADDING', (0,0),(-1,-1), 6),
        ]))
        h2 = Table([[_td(sub_str, size=8, color=WHITE)]],
                   colWidths=[avail_w])
        h2.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), C_BLUE2),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
            ('TOPPADDING',    (0,0),(-1,-1), 3),
            ('BOTTOMPADDING', (0,0),(-1,-1), 3),
        ]))
        return [h1, h2, Spacer(1, 0.25*cm)]

    # ── PAGE 1: Student Marks Table ─────────────────────────────────────────
    story += _page_header('STUDENT MARKS')
    story.append(Paragraph('STUDENT RESULTS TABLE',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.15*cm))
    story.append(_build_marks_table(data, avail_w))

    # ── PAGE 2: Performance Charts ───────────────────────────────────────────
    story.append(PageBreak())
    story += _page_header('PERFORMANCE CHARTS')

    story.append(Paragraph('SUBJECT-WISE AVERAGE',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.1*cm))
    story.append(_subject_average_chart(data, avail_w, 8.5*cm))
    story.append(Spacer(1, 0.35*cm))

    half_w = (avail_w - 0.6*cm) / 2
    gap_tbl = Table([['']], colWidths=[0.6*cm])
    gap_tbl.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0, WHITE)]))
    charts_row = Table(
        [[_grade_distribution_chart(data, half_w, 7.5*cm), gap_tbl,
          _division_distribution_chart(data, half_w, 7.5*cm)]],
        colWidths=[half_w, 0.6*cm, half_w],
    )
    charts_row.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(charts_row)

    # ── PAGE 3: Summary ─────────────────────────────────────────────────────
    story.append(PageBreak())
    story += _page_header('SUMMARY')

    story.append(Paragraph('SUBJECT SUMMARY',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.15*cm))
    story.append(_build_subject_summary_table(data, avail_w))
    story.append(Spacer(1, 0.35*cm))

    story.append(Paragraph('DIVISION SUMMARY',
        _ps('sec', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLUE)))
    story.append(Spacer(1, 0.15*cm))
    story.append(_build_division_block(data, avail_w))
    story.append(Spacer(1, 0.35*cm))

    story.append(_build_best_worst(data, avail_w))

    # ── Footer ──────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.25*cm))
    story.append(HRFlowable(width='100%', thickness=0.4, color=C_MGRAY))
    story.append(Paragraph(
        f'Generated by MathPlatform  ·  {date.today().strftime("%d %B %Y")}  ·  {school_name}',
        _ps('ft', fontSize=6.5, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER),
    ))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def _xl_fill(hex6):
    return PatternFill('solid', fgColor=hex6)

def _xl_font(bold=False, color='FF000000', size=9):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def _xl_align(h='center', v='center', wrap=False, rotate=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, text_rotation=rotate)

def _xl_border(style='thin', color='FFD1D5DB'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_set(ws, row, col, value, bold=False, bg=None, fg='FF000000',
            size=9, align_h='center', wrap=False, number_format=None, rotate=0):
    c = ws.cell(row, col, value)
    c.font      = _xl_font(bold=bold, color=fg, size=size)
    c.alignment = _xl_align(h=align_h, wrap=wrap, rotate=rotate)
    c.border    = _xl_border()
    if bg:
        c.fill = _xl_fill(bg)
    if number_format:
        c.number_format = number_format
    return c

GRADE_BG = {'A':'FF10B981','B':'FF2563EB','C':'FFF59E0B','D':'FFF97316','F':'FFF43F5E'}
HDR1='1E3A5F'; HDR2='2D4F7C'; HDR3='E8F0FE'
GRAY='F3F4F6'; WHITE_XL='FFFFFF'; BLACK_XL='FF000000'


def generate_analytics_report_excel(data: dict, school_name: str = 'MathPlatform') -> bytes:
    wb = Workbook()

    # ── Sheet 1: Student Marks ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Student Marks'
    cls = data['classroom']
    ordered_subjects = data['ordered_subjects']
    n_subj = len(ordered_subjects)

    row = 1
    # Title rows
    ncols_marks = 3 + n_subj * 2 + 4
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_marks)
    _xl_set(ws1, row, 1, f'MathPlatform  ·  {school_name}',
            bold=True, bg=HDR1, fg='FFFFFFFF', size=12)
    ws1.row_dimensions[row].height = 20; row += 1

    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_marks)
    sub = (f"STUDENT RESULTS TABLE  ·  {cls['name']}  ·  "
           f"{cls['academic_year']}  ·  {str(cls['term']).replace('_',' ').upper()}")
    _xl_set(ws1, row, 1, sub, bold=True, bg=HDR2, fg='FFFFFFFF', size=10)
    ws1.row_dimensions[row].height = 16; row += 1
    row += 1  # blank

    # Column headers row
    _xl_set(ws1, row, 1, 'SN',           bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    _xl_set(ws1, row, 2, 'REG NO',       bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    _xl_set(ws1, row, 3, 'STUDENT NAME', bold=True, bg=HDR1, fg='FFFFFFFF', size=8, align_h='left')
    col_i = 4
    for _, smeta in ordered_subjects:
        # Subject code/name is rotated 90° so it fits down a narrow 6-char
        # column instead of forcing the column wide — with many subjects
        # (common in NECTA-scope classes) horizontal labels used to overlap
        # each other and spill into neighbouring score/grade columns.
        label = smeta['code'] or smeta['name'][:10]
        _xl_set(ws1, row, col_i,     label, bold=True, bg=HDR1, fg='FFFFFFFF', size=8, rotate=90)
        _xl_set(ws1, row, col_i + 1, 'GR',  bold=True, bg=HDR2, fg='FFFFFFFF', size=8, rotate=90)
        col_i += 2
    _xl_set(ws1, row, col_i,     'AVG',  bold=True, bg=HDR1, fg='FFFFFFFF', size=8); col_i += 1
    _xl_set(ws1, row, col_i,     'GR',   bold=True, bg=HDR1, fg='FFFFFFFF', size=8); col_i += 1
    _xl_set(ws1, row, col_i,     'DIV',  bold=True, bg=HDR1, fg='FFFFFFFF', size=8); col_i += 1
    _xl_set(ws1, row, col_i,     'POS',  bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    # Tall enough for rotated 10-char subject labels without clipping.
    ws1.row_dimensions[row].height = 70; row += 1

    # Data rows
    for i, sr in enumerate(data['student_rows']):
        bg = GRAY if i % 2 == 0 else WHITE_XL
        _xl_set(ws1, row, 1, i+1,          bg=bg, size=8)
        _xl_set(ws1, row, 2, sr['index'] or '-', bg=bg, size=8)
        _xl_set(ws1, row, 3, sr['name'],   bg=bg, size=8, align_h='left')
        col_i = 4
        for subj_id, _ in ordered_subjects:
            ss  = sr['subj_scores'].get(subj_id, {'pct': None, 'grade': '-'})
            pct = ss['pct']
            g   = ss['grade']
            pct_val = round(pct, 0) if pct is not None else ''
            _xl_set(ws1, row, col_i,     pct_val, bg=bg, size=8,
                    number_format='0' if pct_val != '' else None)
            gbg = GRADE_BG.get(g, WHITE_XL) if g != '-' else bg
            _xl_set(ws1, row, col_i + 1, g, bg=gbg, fg='FFFFFFFF' if g != '-' else BLACK_XL,
                    bold=(g != '-'), size=8)
            col_i += 2
        g_overall = sr['grade']
        _xl_set(ws1, row, col_i,     sr['average'], bg=bg, bold=True, size=8,
                number_format='0.0'); col_i += 1
        _xl_set(ws1, row, col_i,     g_overall,
                bg=GRADE_BG.get(g_overall, WHITE_XL), fg='FFFFFFFF',
                bold=True, size=8); col_i += 1
        _xl_set(ws1, row, col_i,     sr['division'],  bg=bg, size=8); col_i += 1
        _xl_set(ws1, row, col_i,     sr['position'],  bg=bg, size=8)
        ws1.row_dimensions[row].height = 14; row += 1

    # Column widths for sheet 1
    ws1.column_dimensions['A'].width = 4
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 24
    col_letter_i = 4
    for _ in ordered_subjects:
        ws1.column_dimensions[get_column_letter(col_letter_i)].width     = 6
        ws1.column_dimensions[get_column_letter(col_letter_i+1)].width   = 4
        col_letter_i += 2
    for w in [7, 4, 5, 5]:
        ws1.column_dimensions[get_column_letter(col_letter_i)].width = w
        col_letter_i += 1
    ws1.freeze_panes = 'D5'

    # ── Sheet 2: Subject Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    NCOLS = 11
    row = 1

    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 1, f'MathPlatform  ·  {school_name}',
            bold=True, bg=HDR1, fg='FFFFFFFF', size=12)
    ws2.row_dimensions[row].height = 20; row += 1

    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    sub2 = (f"ANALYTICS SUMMARY  ·  {cls['name']}  ·  "
            f"{cls['academic_year']}  ·  {str(cls['term']).replace('_',' ').upper()}")
    _xl_set(ws2, row, 1, sub2, bold=True, bg=HDR2, fg='FFFFFFFF', size=10)
    ws2.row_dimensions[row].height = 16; row += 1
    row += 1

    # Subject summary headers
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 1, 'SUBJECT SUMMARY', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.row_dimensions[row].height = 15; row += 1

    for ci, h in enumerate(['SUBJECT','A','B','C','D','F','SEATS','AVERAGE','GRADE','G.P.A','COMPETENCY'], 1):
        _xl_set(ws2, row, ci, h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    ws2.row_dimensions[row].height = 16; row += 1

    subj_data_start = row
    for i, r in enumerate(data['subject_rows']):
        bg = GRAY if i % 2 == 0 else WHITE_XL
        g  = r['grade']
        _xl_set(ws2, row, 1,  r['subject'],            bg=bg, size=8, align_h='left', bold=True)
        _xl_set(ws2, row, 2,  r['A'],                  bg=bg, size=8)
        _xl_set(ws2, row, 3,  r['B'],                  bg=bg, size=8)
        _xl_set(ws2, row, 4,  r['C'],                  bg=bg, size=8)
        _xl_set(ws2, row, 5,  r['D'],                  bg=bg, size=8)
        _xl_set(ws2, row, 6,  r['F'],                  bg=bg, size=8)
        _xl_set(ws2, row, 7,  r['seats'],              bg=bg, size=8)
        _xl_set(ws2, row, 8,  round(r['average'],2),   bg=bg, size=8, number_format='0.00')
        _xl_set(ws2, row, 9,  g, bg=GRADE_BG.get(g,WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
        _xl_set(ws2, row, 10, round(r['gpa'],4),       bg=bg, size=8, number_format='0.0000')
        _xl_set(ws2, row, 11, r['competency'],         bg=bg, size=8, bold=True)
        ws2.row_dimensions[row].height = 14; row += 1
    subj_data_end = row - 1

    # NECTA-style TOTAL row: A/B/C/D/F summed across every subject entry —
    # the figure real CSEE/FTNA analysis sheets publish as "Grade
    # Distribution". Added below subj_data_end so it doesn't skew the
    # subject-average bar chart, which reads only subj_data_start:subj_data_end.
    ngd = data['necta_grade_distribution']
    total_seats = sum(ngd.values())
    _xl_set(ws2, row, 1, 'TOTAL', bg='D1E8FF', size=8, align_h='left', bold=True)
    _xl_set(ws2, row, 2, ngd['A'], bg='D1E8FF', size=8, bold=True)
    _xl_set(ws2, row, 3, ngd['B'], bg='D1E8FF', size=8, bold=True)
    _xl_set(ws2, row, 4, ngd['C'], bg='D1E8FF', size=8, bold=True)
    _xl_set(ws2, row, 5, ngd['D'], bg='D1E8FF', size=8, bold=True)
    _xl_set(ws2, row, 6, ngd['F'], bg='D1E8FF', size=8, bold=True)
    _xl_set(ws2, row, 7, total_seats, bg='D1E8FF', size=8, bold=True)
    ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
    ws2.row_dimensions[row].height = 14; row += 1

    row += 1
    # Division summary
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    _xl_set(ws2, row, 1, 'DIVISION SUMMARY', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 8, 'CLASSWISE COMPETENCY', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.row_dimensions[row].height = 15; row += 1

    for ci, h in enumerate(['DIVISION','I','II','III','IV','0','N/A'], 1):
        _xl_set(ws2, row, ci, h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    div_hdr_row = row
    cw = data['classwise']
    ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 8,
            f"Overall Subjects Average: {cw['overall_average']}, Grade: {cw['grade']}",
            bg=HDR3, fg='FF1E3A5F', bold=True, size=8, align_h='left')
    ws2.row_dimensions[row].height = 15; row += 1

    # NOTE: StudentProfile has no gender field, so a real M/F split can't be
    # computed — only TOTAL is real data. The first two rows are left blank
    # (rather than fabricated) purely to keep this block the same height as
    # the 3-line classwise text beside it.
    div_summary = data['division_summary']
    cw_extra = [
        f"CLASS G.P.A: {cw['gpa']:.4f}  ({cw['competency']})",
        f"Total Students: {data['total_students']}",
        '',
    ]
    row_labels = ('', '', 'TOTAL')
    div_total_row = None
    for i, label in enumerate(row_labels):
        bg_row = 'D1E8FF' if label == 'TOTAL' else (GRAY if i % 2 == 0 else WHITE_XL)
        bd     = label == 'TOTAL'
        dr     = div_summary.get('TOTAL', {}) if label == 'TOTAL' else {}
        if label == 'TOTAL':
            div_total_row = row
        _xl_set(ws2, row, 1, label,           bg=bg_row, bold=bd, size=8)
        for ci2, div in enumerate(['I','II','III','IV','0'], 2):
            val = dr.get(div, 0) if label == 'TOTAL' else ''
            _xl_set(ws2, row, ci2, val, bg=bg_row, bold=bd, size=8)
        na_val = div_summary.get('INCOMPLETE', 0) if label == 'TOTAL' else ''
        _xl_set(ws2, row, 7, na_val, bg=bg_row, bold=bd, size=8)
        ws2.merge_cells(start_row=row, start_column=8, end_row=row, end_column=NCOLS)
        _xl_set(ws2, row, 8, cw_extra[i],
                bg=HDR3, fg='FF1E3A5F', bold=(i==0), size=8, align_h='left')
        ws2.row_dimensions[row].height = 14; row += 1

    row += 1
    # NECTA-format grade distribution — A/B/C/D/F summed across every
    # subject entry sat by the class (not a per-student average grade),
    # matching how CSEE/FTNA analysis sheets publish it. This table also
    # feeds the pie chart on the Charts sheet.
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _xl_set(ws2, row, 1, 'GRADE DISTRIBUTION (NECTA FORMAT)', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.row_dimensions[row].height = 15; row += 1

    _xl_set(ws2, row, 1, 'GRADE', bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    _xl_set(ws2, row, 2, 'ENTRIES', bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    _xl_set(ws2, row, 3, '%', bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    row += 1
    gd_data_start = row
    ngd_total = sum(data['necta_grade_distribution'].values()) or 1
    for g in ('A', 'B', 'C', 'D', 'F'):
        count = data['necta_grade_distribution'].get(g, 0)
        _xl_set(ws2, row, 1, g, bg=GRADE_BG.get(g, WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
        _xl_set(ws2, row, 2, count, bg=GRAY, size=8)
        _xl_set(ws2, row, 3, count / ngd_total, bg=GRAY, size=8, number_format='0.0%')
        row += 1
    gd_data_end = row - 1

    row += 1
    # Best / Worst
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _xl_set(ws2, row, 1, 'LIST OF 10 BEST STUDENTS', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.merge_cells(start_row=row, start_column=7, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 7, 'LIST OF 10 WORST STUDENTS', bold=True, bg=HDR2, fg='FFFFFFFF', size=9)
    ws2.cell(row, 6).fill = _xl_fill(WHITE_XL)
    ws2.row_dimensions[row].height = 15; row += 1

    sw_hdrs = ['SN','STUDENT NAME','AVERAGE','GRADE','POSITION']
    for ci, h in enumerate(sw_hdrs, 1):
        _xl_set(ws2, row, ci,     h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
        _xl_set(ws2, row, ci + 6, h, bold=True, bg=HDR1, fg='FFFFFFFF', size=8)
    ws2.cell(row, 6).fill = _xl_fill(WHITE_XL)
    ws2.row_dimensions[row].height = 14; row += 1

    max_rows = max(len(data['best_students']), len(data['worst_students']))
    for i in range(max_rows):
        bg = GRAY if i % 2 == 0 else WHITE_XL
        if i < len(data['best_students']):
            br = data['best_students'][i]
            g  = br['grade']
            _xl_set(ws2, row, 1, br['sn'],       bg=bg, size=8)
            _xl_set(ws2, row, 2, br['name'],      bg=bg, size=8, align_h='left')
            _xl_set(ws2, row, 3, br['average'],   bg=bg, size=8, number_format='0.0')
            _xl_set(ws2, row, 4, g, bg=GRADE_BG.get(g,WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
            _xl_set(ws2, row, 5, br['position'],  bg=bg, size=8)
        if i < len(data['worst_students']):
            wr = data['worst_students'][i]
            g  = wr['grade']
            _xl_set(ws2, row, 7,  wr['sn'],       bg=bg, size=8)
            _xl_set(ws2, row, 8,  wr['name'],     bg=bg, size=8, align_h='left')
            _xl_set(ws2, row, 9,  wr['average'],  bg=bg, size=8, number_format='0.0')
            _xl_set(ws2, row, 10, g, bg=GRADE_BG.get(g,WHITE_XL), fg='FFFFFFFF', bold=True, size=8)
            _xl_set(ws2, row, 11, wr['position'], bg=bg, size=8)
        ws2.cell(row, 6).fill = _xl_fill(WHITE_XL)
        ws2.row_dimensions[row].height = 14; row += 1

    # Footer
    row += 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _xl_set(ws2, row, 1,
            f'Generated by MathPlatform  ·  {date.today().strftime("%d %B %Y")}  ·  {school_name}',
            bg=HDR1, fg='FFFFFFFF', size=7)
    ws2.row_dimensions[row].height = 13

    # Column widths sheet 2
    col_w2 = [22, 4.5, 4.5, 4.5, 4.5, 4.5, 6, 10, 7, 9, 16]
    for ci, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = 'A5'

    # ── Sheet 3: Charts (native, data-linked — not static images) ───────────
    ws3 = wb.create_sheet('Charts')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:P1')
    _xl_set(ws3, 1, 1, f'MathPlatform  ·  {school_name}  ·  Performance Charts',
            bold=True, bg=HDR1, fg='FFFFFFFF', size=12)
    ws3.row_dimensions[1].height = 20
    ws3.merge_cells('A2:P2')
    _xl_set(ws3, 2, 1, sub2, bold=True, bg=HDR2, fg='FFFFFFFF', size=10)
    ws3.row_dimensions[2].height = 16

    GRADE_HEX = {k: v[2:] for k, v in GRADE_BG.items()}          # strip 'FF' alpha
    DIVISION_HEX = {'I': '10B981', 'II': '2563EB', 'III': 'F59E0B',
                     'IV': 'F97316', '0': 'F43F5E'}

    def _colour_points(series, colour_map, keys):
        from openpyxl.chart.marker import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties
        series.data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=colour_map.get(k, 'D1D5DB')))
            for i, k in enumerate(keys)
        ]

    # 1) Subject-wise average — bar chart, linked to the Subject Summary table
    bar1 = BarChart()
    bar1.type = 'col'
    bar1.title = 'Subject-wise Class Average (%)'
    bar1.y_axis.title = '% Score'
    bar1.x_axis.title = 'Subject'
    bar1.y_axis.scaling.min = 0
    bar1.y_axis.scaling.max = 100
    bar1.y_axis.numFmt = '0"%"'
    bar1.height = 9
    bar1.width = 18
    bar1.style = 10
    cats1 = Reference(ws2, min_col=1, min_row=subj_data_start, max_row=subj_data_end)
    vals1 = Reference(ws2, min_col=8, min_row=subj_data_start - 1, max_row=subj_data_end)
    bar1.add_data(vals1, titles_from_data=True)
    bar1.set_categories(cats1)
    bar1.dataLabels = DataLabelList()
    bar1.dataLabels.showVal = True
    bar1.dataLabels.showCatName = False
    bar1.dataLabels.showSerName = False
    bar1.dataLabels.showLegendKey = False
    bar1.dataLabels.numFmt = '0.0"%"'
    bar1.legend = None
    if bar1.series:
        _colour_points(bar1.series[0], GRADE_HEX,
                        [r['grade'] for r in data['subject_rows']])
    ws3.add_chart(bar1, 'A4')

    # 2) Class grade distribution — pie chart with grade + % data labels
    pie1 = PieChart()
    pie1.title = 'Grade Distribution (NECTA format)'
    pie1.height = 9
    pie1.width = 12
    cats2 = Reference(ws2, min_col=1, min_row=gd_data_start, max_row=gd_data_end)
    vals2 = Reference(ws2, min_col=2, min_row=gd_data_start - 1, max_row=gd_data_end)
    pie1.add_data(vals2, titles_from_data=True)
    pie1.set_categories(cats2)
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showVal = False
    pie1.dataLabels.showCatName = True
    pie1.dataLabels.showSerName = False
    pie1.dataLabels.showLegendKey = False
    pie1.dataLabels.showPercent = True
    if pie1.series:
        _colour_points(pie1.series[0], GRADE_HEX,
                        [g for g in ('A', 'B', 'C', 'D', 'F')])
    ws3.add_chart(pie1, 'A24')

    # 3) Division distribution — bar chart, linked to the Division Summary row
    bar2 = BarChart()
    bar2.type = 'col'
    bar2.title = 'Division Distribution (Students per Division)'
    bar2.y_axis.title = 'Students'
    bar2.x_axis.title = 'Division'
    bar2.y_axis.numFmt = '0'
    bar2.height = 9
    bar2.width = 12
    cats3 = Reference(ws2, min_col=2, max_col=6, min_row=div_hdr_row)
    vals3 = Reference(ws2, min_col=1, max_col=6, min_row=div_total_row)
    bar2.add_data(vals3, titles_from_data=True, from_rows=True)
    bar2.set_categories(cats3)
    bar2.dataLabels = DataLabelList()
    bar2.dataLabels.showVal = True
    bar2.dataLabels.showCatName = False
    bar2.dataLabels.showSerName = False
    bar2.dataLabels.showLegendKey = False
    bar2.legend = None
    if bar2.series:
        _colour_points(bar2.series[0], DIVISION_HEX, ['I', 'II', 'III', 'IV', '0'])
    ws3.add_chart(bar2, 'K24')

    ws3.column_dimensions['A'].width = 2
    ws3.page_setup.orientation = 'landscape'
    ws3.page_setup.fitToWidth = 1
    ws3.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
